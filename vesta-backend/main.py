import asyncio
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from fastapi import FastAPI, HTTPException, File, UploadFile, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
import csv
import hashlib
import httpx
import io
import json
import math
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
from typing import Any, AsyncGenerator, Dict, List, Literal, Optional, Tuple
from uuid import uuid4

from dotenv import load_dotenv

# Import routing utilities and audit logger
from routing_utils import (
    analyze_message_signals,
    analyze_task_context,
    fast_route,
    enforce_model_consistency,
    should_upgrade_model,
    RoutingDecision,
)
from audit_logger import (
    log_routing_decision,
    log_routing_error,
    log_model_consistency_event,
)

app = FastAPI(title="Vesta Backend")

# CORS middleware for local frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Local only - no auth needed
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Model definitions
DEFAULT_MODEL_NAMES = {
    "general": "hymetalab/vesta-general",
    "deep": "hymetalab/vesta-deep",
    "lite": "hymetalab/vesta-lite",
}
MODEL_PROFILE_KEYS = ("lite", "general", "deep")
ASSISTANT_PROFILE_OPTIONS = {"default", "medical", "legal"}
DEFAULT_ASSISTANT_PROFILE = "default"
EMBEDDING_MODEL = "qwen3-embedding:0.6b"
REQUIRED_VESTA_MODELS = [
    DEFAULT_MODEL_NAMES["lite"],
    DEFAULT_MODEL_NAMES["general"],
    DEFAULT_MODEL_NAMES["deep"],
    EMBEDDING_MODEL,
]

OLLAMA_BASE_URL = "http://localhost:11434"
MAX_CHAT_UPLOAD_SIZE = 10 * 1024 * 1024
MAX_KNOWLEDGE_UPLOAD_SIZE = 25 * 1024 * 1024
MAX_KNOWLEDGE_TEXT_CHARS = 300_000
CHUNK_SIZE = 1_200
CHUNK_OVERLAP = 200
RETRIEVAL_TOP_K = 5
RETRIEVAL_MIN_SCORE = 0.2
FOLDER_COLOR_OPTIONS = {"sand", "stone", "sage", "slate", "taupe", "clay"}
DEFAULT_FOLDER_COLOR = "sand"
DEFAULT_WEATHER_MODE = "general"
WEATHER_MODE_OPTIONS = {
    "storm_damage",
    "lawn_care",
    "construction",
    "general",
}
WEATHER_CACHE_TTL_MINUTES = 45
OPENWEATHER_CURRENT_URL = "https://api.openweathermap.org/data/2.5/weather"
OPENWEATHER_FORECAST_URL = "https://api.openweathermap.org/data/2.5/forecast"
OPENWEATHER_ONECALL_URL = "https://api.openweathermap.org/data/3.0/onecall"
OPENWEATHER_GEOCODE_URL = "https://api.openweathermap.org/geo/1.0/direct"
WEATHER_INTENT_PATTERN = re.compile(
    r"(weather|forecast|storm|rain|snow|wind|hail|humidity|temperature|lawn|construction|alerts?)",
    re.IGNORECASE,
)


def normalize_ollama_model_name(model_name: str) -> str:
    normalized = model_name.strip()
    if not normalized:
        return ""
    if ":" not in normalized:
        return f"{normalized}:latest"
    return normalized


def normalize_assistant_profile(profile: Optional[str]) -> str:
    normalized = (profile or DEFAULT_ASSISTANT_PROFILE).strip().lower()
    if normalized in ASSISTANT_PROFILE_OPTIONS:
        return normalized
    return DEFAULT_ASSISTANT_PROFILE


class ChatMessage(BaseModel):
    role: Literal["user", "assistant"]
    content: str


class ChatRequest(BaseModel):
    mode: Literal["draft", "think", "clarify", "general"]
    profile: Literal["default", "medical", "legal"] = "default"
    message: str = Field(..., min_length=1)
    messages: List[ChatMessage] = Field(default_factory=list)
    model: Optional[Literal["general", "deep", "lite", "auto"]] = "auto"
    last_model_used: Optional[str] = None
    conversation_id: Optional[str] = None
    folder_id: Optional[str] = None


class ChatResponse(BaseModel):
    response: str


class FolderCreateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=120)
    color: Optional[str] = None


class FolderUpdateRequest(BaseModel):
    name: Optional[str] = Field(default=None, min_length=1, max_length=120)
    color: Optional[str] = None


class ConversationCreateRequest(BaseModel):
    folder_id: Optional[str] = None


class ConversationUpdateRequest(BaseModel):
    title: Optional[str] = Field(default=None, min_length=1, max_length=160)
    folder_id: Optional[str] = None


class ConversationTurnRequest(BaseModel):
    user_message: str = Field(..., min_length=1)
    assistant_message: str = Field(..., min_length=1)
    model_used: Optional[str] = None
    sources: Optional[List[Dict[str, Any]]] = None


class ModelSettingsUpdateRequest(BaseModel):
    lite: str = Field(..., min_length=1)
    general: str = Field(..., min_length=1)
    deep: str = Field(..., min_length=1)


class ProfileSettingsUpdateRequest(BaseModel):
    profile: Literal["default", "medical", "legal"]


class SetupPrerequisitesRequest(BaseModel):
    approved: bool = False
    models: Optional[List[str]] = None


class WeatherSettingsUpdateRequest(BaseModel):
    mode: Literal["storm_damage", "lawn_care", "construction", "general"]
    city: str = Field(..., min_length=1, max_length=120)
    state: Optional[str] = Field(default=None, max_length=120)
    country: Optional[str] = Field(default="US", min_length=2, max_length=3)


class KnowledgeStore:
    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self.db_path = data_dir / "knowledge.db"
        self._lock = threading.Lock()
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def _initialize(self) -> None:
        with self._lock:
            with self._connect() as conn:
                conn.executescript(
                    """
                    CREATE TABLE IF NOT EXISTS knowledge_documents (
                        id TEXT PRIMARY KEY,
                        filename TEXT NOT NULL,
                        content_hash TEXT NOT NULL UNIQUE,
                        size_bytes INTEGER NOT NULL,
                        mime_type TEXT,
                        chunk_count INTEGER NOT NULL,
                        created_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS knowledge_chunks (
                        id TEXT PRIMARY KEY,
                        document_id TEXT NOT NULL,
                        chunk_index INTEGER NOT NULL,
                        content TEXT NOT NULL,
                        embedding_json TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY (document_id) REFERENCES knowledge_documents(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_knowledge_chunks_document_id
                        ON knowledge_chunks(document_id);

                    CREATE TABLE IF NOT EXISTS folders (
                        id TEXT PRIMARY KEY,
                        name TEXT NOT NULL UNIQUE,
                        color TEXT NOT NULL DEFAULT 'sand',
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS folder_documents (
                        id TEXT PRIMARY KEY,
                        folder_id TEXT NOT NULL,
                        filename TEXT NOT NULL,
                        content_hash TEXT NOT NULL,
                        size_bytes INTEGER NOT NULL,
                        mime_type TEXT,
                        chunk_count INTEGER NOT NULL,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY (folder_id) REFERENCES folders(id) ON DELETE CASCADE,
                        UNIQUE(folder_id, content_hash)
                    );

                    CREATE TABLE IF NOT EXISTS folder_chunks (
                        id TEXT PRIMARY KEY,
                        document_id TEXT NOT NULL,
                        chunk_index INTEGER NOT NULL,
                        content TEXT NOT NULL,
                        embedding_json TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY (document_id) REFERENCES folder_documents(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_folder_documents_folder_id
                        ON folder_documents(folder_id);

                    CREATE INDEX IF NOT EXISTS idx_folder_chunks_document_id
                        ON folder_chunks(document_id);

                    CREATE TABLE IF NOT EXISTS conversations (
                        id TEXT PRIMARY KEY,
                        title TEXT NOT NULL,
                        folder_id TEXT,
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL,
                        last_message_at TEXT NOT NULL,
                        last_message_preview TEXT NOT NULL,
                        message_count INTEGER NOT NULL DEFAULT 0,
                        FOREIGN KEY (folder_id) REFERENCES folders(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_conversations_folder_id
                        ON conversations(folder_id);

                    CREATE INDEX IF NOT EXISTS idx_conversations_last_message_at
                        ON conversations(last_message_at DESC);

                    CREATE TABLE IF NOT EXISTS conversation_messages (
                        id TEXT PRIMARY KEY,
                        conversation_id TEXT NOT NULL,
                        role TEXT NOT NULL,
                        content TEXT NOT NULL,
                        model_used TEXT,
                        sources_json TEXT,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY (conversation_id) REFERENCES conversations(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_conversation_messages_conversation_id
                        ON conversation_messages(conversation_id, created_at);

                    CREATE TABLE IF NOT EXISTS app_settings (
                        key TEXT PRIMARY KEY,
                        value TEXT NOT NULL,
                        updated_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS setup_runs (
                        id TEXT PRIMARY KEY,
                        requested_models_json TEXT NOT NULL,
                        started_at TEXT NOT NULL,
                        finished_at TEXT,
                        success INTEGER NOT NULL DEFAULT 0,
                        installed_ollama INTEGER NOT NULL DEFAULT 0,
                        started_ollama INTEGER NOT NULL DEFAULT 0,
                        pulled_models_json TEXT NOT NULL DEFAULT '[]',
                        failed_models_json TEXT NOT NULL DEFAULT '[]'
                    );

                    CREATE TABLE IF NOT EXISTS setup_run_events (
                        id TEXT PRIMARY KEY,
                        run_id TEXT NOT NULL,
                        event_type TEXT NOT NULL,
                        message TEXT,
                        model_name TEXT,
                        payload_json TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        FOREIGN KEY (run_id) REFERENCES setup_runs(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_setup_run_events_run_id
                        ON setup_run_events(run_id, created_at);

                    CREATE INDEX IF NOT EXISTS idx_setup_runs_started_at
                        ON setup_runs(started_at DESC);

                    CREATE TABLE IF NOT EXISTS weather_settings (
                        id TEXT PRIMARY KEY,
                        mode TEXT NOT NULL,
                        city TEXT,
                        state TEXT,
                        country TEXT NOT NULL,
                        lat REAL,
                        lon REAL,
                        cache_ttl_minutes INTEGER NOT NULL DEFAULT 45,
                        updated_at TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS weather_data (
                        id TEXT PRIMARY KEY,
                        fetched_at TEXT NOT NULL,
                        lat REAL NOT NULL,
                        lon REAL NOT NULL,
                        temp_f REAL,
                        feels_like_f REAL,
                        humidity_pct REAL,
                        wind_mph REAL,
                        wind_gust_mph REAL,
                        precip_in REAL,
                        condition_code INTEGER,
                        condition_main TEXT,
                        condition_desc TEXT,
                        raw_json TEXT NOT NULL
                    );

                    CREATE INDEX IF NOT EXISTS idx_weather_data_loc_time
                        ON weather_data(lat, lon, fetched_at DESC);

                    CREATE TABLE IF NOT EXISTS forecasts (
                        id TEXT PRIMARY KEY,
                        fetched_at TEXT NOT NULL,
                        forecast_ts TEXT NOT NULL,
                        lat REAL NOT NULL,
                        lon REAL NOT NULL,
                        temp_f REAL,
                        temp_min_f REAL,
                        temp_max_f REAL,
                        humidity_pct REAL,
                        wind_mph REAL,
                        precip_prob REAL,
                        precip_in REAL,
                        condition_code INTEGER,
                        condition_main TEXT,
                        condition_desc TEXT,
                        confidence_score REAL,
                        raw_json TEXT NOT NULL
                    );

                    CREATE INDEX IF NOT EXISTS idx_forecasts_loc_time
                        ON forecasts(lat, lon, forecast_ts);

                    CREATE TABLE IF NOT EXISTS weather_alerts (
                        id TEXT PRIMARY KEY,
                        fetched_at TEXT NOT NULL,
                        lat REAL NOT NULL,
                        lon REAL NOT NULL,
                        event TEXT,
                        severity TEXT,
                        sender_name TEXT,
                        start_ts TEXT,
                        end_ts TEXT,
                        description TEXT,
                        raw_json TEXT NOT NULL
                    );

                    CREATE TABLE IF NOT EXISTS predictions (
                        id TEXT PRIMARY KEY,
                        created_at TEXT NOT NULL,
                        mode TEXT NOT NULL,
                        target_date TEXT NOT NULL,
                        integrity REAL NOT NULL,
                        resilience REAL NOT NULL,
                        meaning REAL NOT NULL,
                        cci_score REAL NOT NULL,
                        probability REAL NOT NULL,
                        prediction_text TEXT NOT NULL,
                        factors_json TEXT NOT NULL
                    );

                    CREATE INDEX IF NOT EXISTS idx_predictions_mode_target
                        ON predictions(mode, target_date);

                    CREATE TABLE IF NOT EXISTS prediction_outcomes (
                        id TEXT PRIMARY KEY,
                        prediction_id TEXT NOT NULL,
                        evaluated_at TEXT NOT NULL,
                        actual_outcome INTEGER NOT NULL,
                        accuracy_score REAL NOT NULL,
                        notes TEXT,
                        FOREIGN KEY (prediction_id) REFERENCES predictions(id) ON DELETE CASCADE
                    );

                    CREATE INDEX IF NOT EXISTS idx_prediction_outcomes_prediction_id
                        ON prediction_outcomes(prediction_id);

                    CREATE TABLE IF NOT EXISTS weather_api_log (
                        id TEXT PRIMARY KEY,
                        timestamp TEXT NOT NULL,
                        endpoint TEXT NOT NULL,
                        success INTEGER NOT NULL,
                        http_status INTEGER,
                        latency_ms REAL,
                        error_text TEXT
                    );

                    CREATE INDEX IF NOT EXISTS idx_weather_api_log_time
                        ON weather_api_log(timestamp DESC);
                    """
                )
                columns = {
                    row["name"]
                    for row in conn.execute("PRAGMA table_info(folders)").fetchall()
                }
                if "color" not in columns:
                    conn.execute(
                        f"""
                        ALTER TABLE folders
                        ADD COLUMN color TEXT NOT NULL DEFAULT '{DEFAULT_FOLDER_COLOR}'
                        """
                    )
                conn.execute(
                    """
                    UPDATE folders
                    SET color = ?
                    WHERE color IS NULL OR TRIM(color) = ''
                    """,
                    (DEFAULT_FOLDER_COLOR,),
                )
                now = self._current_ts()
                for profile_key in MODEL_PROFILE_KEYS:
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO app_settings (key, value, updated_at)
                        VALUES (?, ?, ?)
                        """,
                        (profile_key, DEFAULT_MODEL_NAMES[profile_key], now),
                    )
                conn.execute(
                    """
                    INSERT OR IGNORE INTO app_settings (key, value, updated_at)
                    VALUES ('assistant_profile', ?, ?)
                    """,
                    (DEFAULT_ASSISTANT_PROFILE, now),
                )
                conn.execute(
                    """
                    INSERT OR IGNORE INTO weather_settings (
                        id, mode, country, cache_ttl_minutes, updated_at
                    )
                    VALUES ('default', ?, 'US', ?, ?)
                    """,
                    (DEFAULT_WEATHER_MODE, WEATHER_CACHE_TTL_MINUTES, now),
                )
                conn.commit()

    def _current_ts(self) -> str:
        return str(int(time.time()))

    def get_model_names(self) -> Dict[str, str]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT key, value
                    FROM app_settings
                    WHERE key IN (?, ?, ?)
                    """,
                    MODEL_PROFILE_KEYS,
                ).fetchall()

                configured = dict(DEFAULT_MODEL_NAMES)
                for row in rows:
                    key = str(row["key"])
                    value = str(row["value"]).strip()
                    if key in MODEL_PROFILE_KEYS and value:
                        configured[key] = value

                return configured

    def set_model_names(self, *, lite: str, general: str, deep: str) -> Dict[str, str]:
        now = self._current_ts()
        next_config = {
            "lite": lite.strip(),
            "general": general.strip(),
            "deep": deep.strip(),
        }

        with self._lock:
            with self._connect() as conn:
                conn.executemany(
                    """
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES (?, ?, ?)
                    ON CONFLICT(key) DO UPDATE SET
                        value = excluded.value,
                        updated_at = excluded.updated_at
                    """,
                    [
                        (profile_key, next_config[profile_key], now)
                        for profile_key in MODEL_PROFILE_KEYS
                    ],
                )
                conn.commit()

        return next_config

    def get_assistant_profile(self) -> str:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT value
                    FROM app_settings
                    WHERE key = 'assistant_profile'
                    """
                ).fetchone()
                if row is None:
                    return DEFAULT_ASSISTANT_PROFILE

                return normalize_assistant_profile(str(row["value"]))

    def set_assistant_profile(self, profile: str) -> str:
        normalized = normalize_assistant_profile(profile)
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES ('assistant_profile', ?, ?)
                    ON CONFLICT(key) DO UPDATE SET
                        value = excluded.value,
                        updated_at = excluded.updated_at
                    """,
                    (normalized, now),
                )
                conn.commit()

        return normalized

    def create_setup_run(self, requested_models: Optional[List[str]]) -> str:
        run_id = str(uuid4())
        now = self._current_ts()
        payload = [model.strip() for model in (requested_models or []) if model.strip()]

        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO setup_runs (
                        id,
                        requested_models_json,
                        started_at,
                        success,
                        installed_ollama,
                        started_ollama,
                        pulled_models_json,
                        failed_models_json
                    ) VALUES (?, ?, ?, 0, 0, 0, '[]', '[]')
                    """,
                    (run_id, json.dumps(payload), now),
                )
                conn.commit()

        return run_id

    def append_setup_run_event(
        self,
        run_id: str,
        event_type: str,
        *,
        message: Optional[str] = None,
        model_name: Optional[str] = None,
        payload: Optional[Dict[str, Any]] = None,
    ) -> None:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO setup_run_events (
                        id, run_id, event_type, message, model_name, payload_json, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        run_id,
                        event_type,
                        message,
                        model_name,
                        json.dumps(payload or {}),
                        now,
                    ),
                )
                conn.commit()

    def finish_setup_run(
        self,
        run_id: str,
        *,
        success: bool,
        installed_ollama: bool,
        started_ollama: bool,
        pulled_models: List[str],
        failed_models: List[Dict[str, str]],
    ) -> None:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    UPDATE setup_runs
                    SET
                        finished_at = ?,
                        success = ?,
                        installed_ollama = ?,
                        started_ollama = ?,
                        pulled_models_json = ?,
                        failed_models_json = ?
                    WHERE id = ?
                    """,
                    (
                        now,
                        1 if success else 0,
                        1 if installed_ollama else 0,
                        1 if started_ollama else 0,
                        json.dumps(pulled_models),
                        json.dumps(failed_models),
                        run_id,
                    ),
                )
                conn.commit()

    def list_setup_runs(self, limit: int = 20) -> List[Dict[str, Any]]:
        safe_limit = max(1, min(limit, 200))
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        id,
                        requested_models_json,
                        started_at,
                        finished_at,
                        success,
                        installed_ollama,
                        started_ollama,
                        pulled_models_json,
                        failed_models_json
                    FROM setup_runs
                    ORDER BY started_at DESC
                    LIMIT ?
                    """,
                    (safe_limit,),
                ).fetchall()

                runs: List[Dict[str, Any]] = []
                for row in rows:
                    run_data = dict(row)
                    run_id = str(run_data["id"])
                    run_data["requested_models"] = json.loads(
                        str(run_data.pop("requested_models_json") or "[]")
                    )
                    run_data["pulled_models"] = json.loads(
                        str(run_data.pop("pulled_models_json") or "[]")
                    )
                    run_data["failed_models"] = json.loads(
                        str(run_data.pop("failed_models_json") or "[]")
                    )
                    run_data["success"] = bool(run_data["success"])
                    run_data["installed_ollama"] = bool(run_data["installed_ollama"])
                    run_data["started_ollama"] = bool(run_data["started_ollama"])

                    event_rows = conn.execute(
                        """
                        SELECT event_type, message, model_name, payload_json, created_at
                        FROM setup_run_events
                        WHERE run_id = ?
                        ORDER BY created_at ASC
                        """,
                        (run_id,),
                    ).fetchall()
                    events: List[Dict[str, Any]] = []
                    for event_row in event_rows:
                        event_data = dict(event_row)
                        event_data["payload"] = json.loads(
                            str(event_data.pop("payload_json") or "{}")
                        )
                        events.append(event_data)

                    run_data["events"] = events
                    runs.append(run_data)

                return runs

    def get_weather_settings(self) -> Dict[str, Any]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT id, mode, city, state, country, lat, lon, cache_ttl_minutes, updated_at
                    FROM weather_settings
                    WHERE id = 'default'
                    """
                ).fetchone()
                if row is None:
                    now = self._current_ts()
                    conn.execute(
                        """
                        INSERT INTO weather_settings (
                            id, mode, country, cache_ttl_minutes, updated_at
                        ) VALUES ('default', ?, 'US', ?, ?)
                        """,
                        (DEFAULT_WEATHER_MODE, WEATHER_CACHE_TTL_MINUTES, now),
                    )
                    conn.commit()
                    row = conn.execute(
                        """
                        SELECT id, mode, city, state, country, lat, lon, cache_ttl_minutes, updated_at
                        FROM weather_settings
                        WHERE id = 'default'
                        """
                    ).fetchone()

                if row is None:
                    return {
                        "mode": DEFAULT_WEATHER_MODE,
                        "location": None,
                        "cache_ttl_minutes": WEATHER_CACHE_TTL_MINUTES,
                        "updated_at": self._current_ts(),
                    }

                data = dict(row)
                location = None
                if data.get("lat") is not None and data.get("lon") is not None and data.get("city"):
                    location = {
                        "city": data["city"],
                        "state": data.get("state"),
                        "country": data.get("country") or "US",
                        "lat": float(data["lat"]),
                        "lon": float(data["lon"]),
                    }

                return {
                    "mode": str(data.get("mode") or DEFAULT_WEATHER_MODE),
                    "location": location,
                    "cache_ttl_minutes": int(data.get("cache_ttl_minutes") or WEATHER_CACHE_TTL_MINUTES),
                    "updated_at": str(data.get("updated_at") or self._current_ts()),
                }

    def set_weather_settings(
        self,
        *,
        mode: str,
        city: str,
        state: Optional[str],
        country: str,
        lat: float,
        lon: float,
        cache_ttl_minutes: int = WEATHER_CACHE_TTL_MINUTES,
    ) -> Dict[str, Any]:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO weather_settings (
                        id, mode, city, state, country, lat, lon, cache_ttl_minutes, updated_at
                    ) VALUES ('default', ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                        mode = excluded.mode,
                        city = excluded.city,
                        state = excluded.state,
                        country = excluded.country,
                        lat = excluded.lat,
                        lon = excluded.lon,
                        cache_ttl_minutes = excluded.cache_ttl_minutes,
                        updated_at = excluded.updated_at
                    """,
                    (
                        mode.strip(),
                        city.strip(),
                        state.strip() if state else None,
                        country.strip().upper(),
                        float(lat),
                        float(lon),
                        int(cache_ttl_minutes),
                        now,
                    ),
                )
                conn.commit()

        return self.get_weather_settings()

    def has_cached_weather_data(self) -> bool:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute("SELECT COUNT(*) AS count FROM weather_data").fetchone()
                return int(row["count"]) > 0 if row else False

    def get_last_weather_refresh_ts(self) -> Optional[str]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT MAX(fetched_at) AS latest FROM weather_data"
                ).fetchone()
                if not row:
                    return None
                latest = row["latest"]
                return str(latest) if latest else None

    def get_latest_weather_current(self, lat: float, lon: float) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT *
                    FROM weather_data
                    WHERE ABS(lat - ?) < 0.01 AND ABS(lon - ?) < 0.01
                    ORDER BY fetched_at DESC
                    LIMIT 1
                    """,
                    (lat, lon),
                ).fetchone()
                return dict(row) if row else None

    def get_latest_weather_forecast(self, lat: float, lon: float) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                fetched_row = conn.execute(
                    """
                    SELECT fetched_at
                    FROM forecasts
                    WHERE ABS(lat - ?) < 0.01 AND ABS(lon - ?) < 0.01
                    ORDER BY fetched_at DESC
                    LIMIT 1
                    """,
                    (lat, lon),
                ).fetchone()
                if not fetched_row:
                    return []

                fetched_at = str(fetched_row["fetched_at"])
                rows = conn.execute(
                    """
                    SELECT *
                    FROM forecasts
                    WHERE fetched_at = ? AND ABS(lat - ?) < 0.01 AND ABS(lon - ?) < 0.01
                    ORDER BY forecast_ts ASC
                    """,
                    (fetched_at, lat, lon),
                ).fetchall()
                return [dict(row) for row in rows]

    def get_latest_weather_alerts(self, lat: float, lon: float) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                fetched_row = conn.execute(
                    """
                    SELECT fetched_at
                    FROM weather_alerts
                    WHERE ABS(lat - ?) < 0.01 AND ABS(lon - ?) < 0.01
                    ORDER BY fetched_at DESC
                    LIMIT 1
                    """,
                    (lat, lon),
                ).fetchone()
                if not fetched_row:
                    return []

                fetched_at = str(fetched_row["fetched_at"])
                rows = conn.execute(
                    """
                    SELECT *
                    FROM weather_alerts
                    WHERE fetched_at = ? AND ABS(lat - ?) < 0.01 AND ABS(lon - ?) < 0.01
                    ORDER BY start_ts ASC
                    """,
                    (fetched_at, lat, lon),
                ).fetchall()
                return [dict(row) for row in rows]

    def save_weather_snapshot(
        self,
        *,
        lat: float,
        lon: float,
        current: Dict[str, Any],
        forecast_points: List[Dict[str, Any]],
        alerts: List[Dict[str, Any]],
        fetched_at: Optional[str] = None,
    ) -> str:
        target_fetched_at = fetched_at or self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO weather_data (
                        id, fetched_at, lat, lon, temp_f, feels_like_f, humidity_pct,
                        wind_mph, wind_gust_mph, precip_in, condition_code, condition_main,
                        condition_desc, raw_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        target_fetched_at,
                        float(lat),
                        float(lon),
                        current.get("temp_f"),
                        current.get("feels_like_f"),
                        current.get("humidity_pct"),
                        current.get("wind_mph"),
                        current.get("wind_gust_mph"),
                        current.get("precip_in"),
                        current.get("condition_code"),
                        current.get("condition_main"),
                        current.get("condition_desc"),
                        json.dumps(current.get("raw_json", {})),
                    ),
                )

                for point in forecast_points:
                    conn.execute(
                        """
                        INSERT INTO forecasts (
                            id, fetched_at, forecast_ts, lat, lon, temp_f, temp_min_f, temp_max_f,
                            humidity_pct, wind_mph, precip_prob, precip_in, condition_code,
                            condition_main, condition_desc, confidence_score, raw_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            target_fetched_at,
                            point.get("forecast_ts"),
                            float(lat),
                            float(lon),
                            point.get("temp_f"),
                            point.get("temp_min_f"),
                            point.get("temp_max_f"),
                            point.get("humidity_pct"),
                            point.get("wind_mph"),
                            point.get("precip_prob"),
                            point.get("precip_in"),
                            point.get("condition_code"),
                            point.get("condition_main"),
                            point.get("condition_desc"),
                            point.get("confidence_score"),
                            json.dumps(point.get("raw_json", {})),
                        ),
                    )

                for alert in alerts:
                    conn.execute(
                        """
                        INSERT INTO weather_alerts (
                            id, fetched_at, lat, lon, event, severity, sender_name,
                            start_ts, end_ts, description, raw_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            target_fetched_at,
                            float(lat),
                            float(lon),
                            alert.get("event"),
                            alert.get("severity"),
                            alert.get("sender_name"),
                            alert.get("start_ts"),
                            alert.get("end_ts"),
                            alert.get("description"),
                            json.dumps(alert.get("raw_json", {})),
                        ),
                    )

                conn.commit()
        return target_fetched_at

    def save_current_weather(
        self,
        *,
        lat: float,
        lon: float,
        current: Dict[str, Any],
        fetched_at: Optional[str] = None,
    ) -> str:
        target_fetched_at = fetched_at or self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO weather_data (
                        id, fetched_at, lat, lon, temp_f, feels_like_f, humidity_pct,
                        wind_mph, wind_gust_mph, precip_in, condition_code, condition_main,
                        condition_desc, raw_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        target_fetched_at,
                        float(lat),
                        float(lon),
                        current.get("temp_f"),
                        current.get("feels_like_f"),
                        current.get("humidity_pct"),
                        current.get("wind_mph"),
                        current.get("wind_gust_mph"),
                        current.get("precip_in"),
                        current.get("condition_code"),
                        current.get("condition_main"),
                        current.get("condition_desc"),
                        json.dumps(current.get("raw_json", {})),
                    ),
                )
                conn.commit()
        return target_fetched_at

    def save_forecast_weather(
        self,
        *,
        lat: float,
        lon: float,
        forecast_points: List[Dict[str, Any]],
        fetched_at: Optional[str] = None,
    ) -> str:
        target_fetched_at = fetched_at or self._current_ts()
        with self._lock:
            with self._connect() as conn:
                for point in forecast_points:
                    conn.execute(
                        """
                        INSERT INTO forecasts (
                            id, fetched_at, forecast_ts, lat, lon, temp_f, temp_min_f, temp_max_f,
                            humidity_pct, wind_mph, precip_prob, precip_in, condition_code,
                            condition_main, condition_desc, confidence_score, raw_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            target_fetched_at,
                            point.get("forecast_ts"),
                            float(lat),
                            float(lon),
                            point.get("temp_f"),
                            point.get("temp_min_f"),
                            point.get("temp_max_f"),
                            point.get("humidity_pct"),
                            point.get("wind_mph"),
                            point.get("precip_prob"),
                            point.get("precip_in"),
                            point.get("condition_code"),
                            point.get("condition_main"),
                            point.get("condition_desc"),
                            point.get("confidence_score"),
                            json.dumps(point.get("raw_json", {})),
                        ),
                    )
                conn.commit()
        return target_fetched_at

    def save_weather_alerts(
        self,
        *,
        lat: float,
        lon: float,
        alerts: List[Dict[str, Any]],
        fetched_at: Optional[str] = None,
    ) -> str:
        target_fetched_at = fetched_at or self._current_ts()
        with self._lock:
            with self._connect() as conn:
                for alert in alerts:
                    conn.execute(
                        """
                        INSERT INTO weather_alerts (
                            id, fetched_at, lat, lon, event, severity, sender_name,
                            start_ts, end_ts, description, raw_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            target_fetched_at,
                            float(lat),
                            float(lon),
                            alert.get("event"),
                            alert.get("severity"),
                            alert.get("sender_name"),
                            alert.get("start_ts"),
                            alert.get("end_ts"),
                            alert.get("description"),
                            json.dumps(alert.get("raw_json", {})),
                        ),
                    )
                conn.commit()
        return target_fetched_at

    def log_weather_api_call(
        self,
        *,
        endpoint: str,
        success: bool,
        http_status: Optional[int],
        latency_ms: Optional[float],
        error_text: Optional[str] = None,
    ) -> None:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO weather_api_log (
                        id, timestamp, endpoint, success, http_status, latency_ms, error_text
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        now,
                        endpoint,
                        1 if success else 0,
                        http_status,
                        latency_ms,
                        error_text,
                    ),
                )
                conn.commit()

    def get_weather_api_reliability(self, sample_size: int = 120) -> float:
        safe_size = max(1, min(sample_size, 500))
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT success
                    FROM weather_api_log
                    ORDER BY timestamp DESC
                    LIMIT ?
                    """,
                    (safe_size,),
                ).fetchall()
                if not rows:
                    return 1.0

                success_total = sum(int(row["success"]) for row in rows)
                return success_total / max(len(rows), 1)

    def get_latest_weather_api_failure(self) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT endpoint, http_status, error_text, timestamp
                    FROM weather_api_log
                    WHERE success = 0
                    ORDER BY timestamp DESC
                    LIMIT 1
                    """
                ).fetchone()
                return dict(row) if row else None

    def get_latest_weather_api_event(self) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT endpoint, success, http_status, error_text, timestamp
                    FROM weather_api_log
                    ORDER BY timestamp DESC
                    LIMIT 1
                    """
                ).fetchone()
                return dict(row) if row else None

    def save_predictions(
        self,
        *,
        mode: str,
        integrity: float,
        resilience: float,
        meaning: float,
        cci_score: float,
        predictions: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        now = self._current_ts()
        saved: List[Dict[str, Any]] = []
        with self._lock:
            with self._connect() as conn:
                for item in predictions:
                    prediction_id = str(uuid4())
                    factors = item.get("factors") or {}
                    conn.execute(
                        """
                        INSERT INTO predictions (
                            id, created_at, mode, target_date, integrity, resilience, meaning,
                            cci_score, probability, prediction_text, factors_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            prediction_id,
                            now,
                            mode,
                            item["target_date"],
                            integrity,
                            resilience,
                            meaning,
                            cci_score,
                            item["probability"],
                            item["prediction_text"],
                            json.dumps(factors),
                        ),
                    )
                    saved.append(
                        {
                            "id": prediction_id,
                            "created_at": now,
                            "mode": mode,
                            "target_date": item["target_date"],
                            "integrity": integrity,
                            "resilience": resilience,
                            "meaning": meaning,
                            "cci_score": cci_score,
                            "probability": item["probability"],
                            "prediction_text": item["prediction_text"],
                            "factors": factors,
                        }
                    )
                conn.commit()
        return saved

    def list_predictions(
        self,
        *,
        mode: str,
        date_from: str,
        limit: int = 20,
    ) -> List[Dict[str, Any]]:
        safe_limit = max(1, min(limit, 100))
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        p.id,
                        p.created_at,
                        p.mode,
                        p.target_date,
                        p.integrity,
                        p.resilience,
                        p.meaning,
                        p.cci_score,
                        p.probability,
                        p.prediction_text,
                        p.factors_json,
                        o.actual_outcome,
                        o.accuracy_score,
                        o.evaluated_at
                    FROM predictions p
                    LEFT JOIN prediction_outcomes o ON o.prediction_id = p.id
                    WHERE p.mode = ? AND p.target_date >= ?
                    ORDER BY p.target_date ASC, p.created_at DESC
                    LIMIT ?
                    """,
                    (mode, date_from, safe_limit),
                ).fetchall()

                data: List[Dict[str, Any]] = []
                for row in rows:
                    item = dict(row)
                    try:
                        item["factors"] = json.loads(item.pop("factors_json") or "{}")
                    except Exception:
                        item["factors"] = {}
                    data.append(item)
                return data

    def list_mature_predictions_without_outcome(self, date_before: str) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT p.id, p.mode, p.target_date, p.probability, p.factors_json
                    FROM predictions p
                    LEFT JOIN prediction_outcomes o ON o.prediction_id = p.id
                    WHERE o.id IS NULL AND p.target_date < ?
                    ORDER BY p.target_date ASC
                    """,
                    (date_before,),
                ).fetchall()
                results: List[Dict[str, Any]] = []
                for row in rows:
                    item = dict(row)
                    try:
                        item["factors"] = json.loads(item.pop("factors_json") or "{}")
                    except Exception:
                        item["factors"] = {}
                    results.append(item)
                return results

    def save_prediction_outcome(
        self,
        *,
        prediction_id: str,
        actual_outcome: bool,
        accuracy_score: float,
        notes: Optional[str] = None,
    ) -> None:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO prediction_outcomes (
                        id, prediction_id, evaluated_at, actual_outcome, accuracy_score, notes
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        prediction_id,
                        now,
                        1 if actual_outcome else 0,
                        accuracy_score,
                        notes,
                    ),
                )
                conn.commit()

    def list_prediction_accuracy(self, mode: str) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT p.factors_json, o.accuracy_score
                    FROM prediction_outcomes o
                    JOIN predictions p ON p.id = o.prediction_id
                    WHERE p.mode = ?
                    ORDER BY o.evaluated_at DESC
                    LIMIT 300
                    """,
                    (mode,),
                ).fetchall()
                results: List[Dict[str, Any]] = []
                for row in rows:
                    item = dict(row)
                    try:
                        item["factors"] = json.loads(item.pop("factors_json") or "{}")
                    except Exception:
                        item["factors"] = {}
                    results.append(item)
                return results

    def get_document_by_hash(self, content_hash: str) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    FROM knowledge_documents
                    WHERE content_hash = ?
                    """,
                    (content_hash,),
                ).fetchone()
                return dict(row) if row else None

    def get_folder_document_by_hash(
        self, folder_id: str, content_hash: str
    ) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT id, folder_id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    FROM folder_documents
                    WHERE folder_id = ? AND content_hash = ?
                    """,
                    (folder_id, content_hash),
                ).fetchone()
                return dict(row) if row else None

    def insert_document_with_chunks(
        self,
        *,
        filename: str,
        content_hash: str,
        size_bytes: int,
        mime_type: Optional[str],
        chunks: List[str],
        embeddings: List[List[float]],
    ) -> Dict[str, Any]:
        if len(chunks) != len(embeddings):
            raise ValueError("Chunks and embeddings length mismatch")

        now = self._current_ts()
        document_id = str(uuid4())

        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO knowledge_documents (
                        id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        document_id,
                        filename,
                        content_hash,
                        size_bytes,
                        mime_type,
                        len(chunks),
                        now,
                    ),
                )

                for index, (chunk_text, embedding) in enumerate(zip(chunks, embeddings)):
                    conn.execute(
                        """
                        INSERT INTO knowledge_chunks (
                            id, document_id, chunk_index, content, embedding_json, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            document_id,
                            index,
                            chunk_text,
                            json.dumps(embedding),
                            now,
                        ),
                    )

                conn.commit()

        return {
            "id": document_id,
            "filename": filename,
            "content_hash": content_hash,
            "size_bytes": size_bytes,
            "mime_type": mime_type,
            "chunk_count": len(chunks),
            "created_at": now,
        }

    def insert_folder_document_with_chunks(
        self,
        *,
        folder_id: str,
        filename: str,
        content_hash: str,
        size_bytes: int,
        mime_type: Optional[str],
        chunks: List[str],
        embeddings: List[List[float]],
    ) -> Dict[str, Any]:
        if len(chunks) != len(embeddings):
            raise ValueError("Chunks and embeddings length mismatch")

        now = self._current_ts()
        document_id = str(uuid4())

        with self._lock:
            with self._connect() as conn:
                folder = conn.execute(
                    "SELECT id FROM folders WHERE id = ?",
                    (folder_id,),
                ).fetchone()
                if folder is None:
                    raise ValueError("Folder not found")

                conn.execute(
                    """
                    INSERT INTO folder_documents (
                        id, folder_id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        document_id,
                        folder_id,
                        filename,
                        content_hash,
                        size_bytes,
                        mime_type,
                        len(chunks),
                        now,
                    ),
                )

                for index, (chunk_text, embedding) in enumerate(zip(chunks, embeddings)):
                    conn.execute(
                        """
                        INSERT INTO folder_chunks (
                            id, document_id, chunk_index, content, embedding_json, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            str(uuid4()),
                            document_id,
                            index,
                            chunk_text,
                            json.dumps(embedding),
                            now,
                        ),
                    )

                conn.commit()

        return {
            "id": document_id,
            "folder_id": folder_id,
            "filename": filename,
            "content_hash": content_hash,
            "size_bytes": size_bytes,
            "mime_type": mime_type,
            "chunk_count": len(chunks),
            "created_at": now,
        }

    def list_documents(self) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    FROM knowledge_documents
                    ORDER BY created_at DESC
                    """
                ).fetchall()
                return [dict(row) for row in rows]

    def list_folder_documents(self, folder_id: str) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT id, folder_id, filename, content_hash, size_bytes, mime_type, chunk_count, created_at
                    FROM folder_documents
                    WHERE folder_id = ?
                    ORDER BY created_at DESC
                    """,
                    (folder_id,),
                ).fetchall()
                return [dict(row) for row in rows]

    def delete_document(self, document_id: str) -> Dict[str, Any]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT chunk_count FROM knowledge_documents WHERE id = ?",
                    (document_id,),
                ).fetchone()
                if row is None:
                    return {"deleted": False, "chunk_count": 0}

                chunk_count = int(row["chunk_count"])
                conn.execute("DELETE FROM knowledge_chunks WHERE document_id = ?", (document_id,))
                conn.execute("DELETE FROM knowledge_documents WHERE id = ?", (document_id,))
                conn.commit()
                return {"deleted": True, "chunk_count": chunk_count}

    def delete_folder_document(self, folder_id: str, document_id: str) -> Dict[str, Any]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT chunk_count
                    FROM folder_documents
                    WHERE id = ? AND folder_id = ?
                    """,
                    (document_id, folder_id),
                ).fetchone()
                if row is None:
                    return {"deleted": False, "chunk_count": 0}

                chunk_count = int(row["chunk_count"])
                conn.execute("DELETE FROM folder_chunks WHERE document_id = ?", (document_id,))
                conn.execute(
                    "DELETE FROM folder_documents WHERE id = ? AND folder_id = ?",
                    (document_id, folder_id),
                )
                conn.commit()
                return {"deleted": True, "chunk_count": chunk_count}

    def get_all_chunks(self) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        c.id,
                        c.document_id,
                        c.chunk_index,
                        c.content,
                        c.embedding_json,
                        d.filename
                    FROM knowledge_chunks c
                    JOIN knowledge_documents d ON c.document_id = d.id
                    """
                ).fetchall()
                return [dict(row) for row in rows]

    def get_folder_chunks(self, folder_id: str) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        c.id,
                        d.id AS document_id,
                        d.folder_id AS folder_id,
                        f.name AS folder_name,
                        c.chunk_index,
                        c.content,
                        c.embedding_json,
                        d.filename
                    FROM folder_chunks c
                    JOIN folder_documents d ON c.document_id = d.id
                    JOIN folders f ON d.folder_id = f.id
                    WHERE d.folder_id = ?
                    """,
                    (folder_id,),
                ).fetchall()
                return [dict(row) for row in rows]

    def list_folders(self) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        f.id,
                        f.name,
                        f.color,
                        f.created_at,
                        f.updated_at,
                        COALESCE(cc.chat_count, 0) AS chat_count,
                        COALESCE(dc.document_count, 0) AS document_count
                    FROM folders f
                    LEFT JOIN (
                        SELECT folder_id, COUNT(*) AS chat_count
                        FROM conversations
                        WHERE folder_id IS NOT NULL
                        GROUP BY folder_id
                    ) cc ON cc.folder_id = f.id
                    LEFT JOIN (
                        SELECT folder_id, COUNT(*) AS document_count
                        FROM folder_documents
                        GROUP BY folder_id
                    ) dc ON dc.folder_id = f.id
                    ORDER BY f.updated_at DESC, f.created_at DESC
                    """
                ).fetchall()
                return [dict(row) for row in rows]

    def get_folder(self, folder_id: str) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    """
                    SELECT id, name, color, created_at, updated_at
                    FROM folders
                    WHERE id = ?
                    """,
                    (folder_id,),
                ).fetchone()
                return dict(row) if row else None

    def create_folder(self, name: str, color: str) -> Dict[str, Any]:
        now = self._current_ts()
        folder_id = str(uuid4())

        with self._lock:
            with self._connect() as conn:
                conn.execute(
                    """
                    INSERT INTO folders (id, name, color, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (folder_id, name.strip(), color, now, now),
                )
                conn.commit()

        return {
            "id": folder_id,
            "name": name.strip(),
            "color": color,
            "created_at": now,
            "updated_at": now,
            "chat_count": 0,
            "document_count": 0,
        }

    def update_folder(
        self,
        folder_id: str,
        *,
        name: Optional[str] = None,
        color: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        now = self._current_ts()
        with self._lock:
            with self._connect() as conn:
                assignments: List[str] = []
                params: List[Any] = []

                if name is not None:
                    assignments.append("name = ?")
                    params.append(name.strip())

                if color is not None:
                    assignments.append("color = ?")
                    params.append(color)

                assignments.append("updated_at = ?")
                params.append(now)
                params.append(folder_id)

                result = conn.execute(
                    f"""
                    UPDATE folders
                    SET {", ".join(assignments)}
                    WHERE id = ?
                    """,
                    tuple(params),
                )
                conn.commit()
                if result.rowcount == 0:
                    return None

                row = conn.execute(
                    """
                    SELECT id, name, color, created_at, updated_at
                    FROM folders
                    WHERE id = ?
                    """,
                    (folder_id,),
                ).fetchone()
                if row is None:
                    return None

                chat_count = conn.execute(
                    "SELECT COUNT(*) AS count FROM conversations WHERE folder_id = ?",
                    (folder_id,),
                ).fetchone()["count"]
                document_count = conn.execute(
                    "SELECT COUNT(*) AS count FROM folder_documents WHERE folder_id = ?",
                    (folder_id,),
                ).fetchone()["count"]

                data = dict(row)
                data["chat_count"] = int(chat_count)
                data["document_count"] = int(document_count)
                return data

    def delete_folder(self, folder_id: str) -> Dict[str, Any]:
        with self._lock:
            with self._connect() as conn:
                exists = conn.execute(
                    "SELECT id FROM folders WHERE id = ?",
                    (folder_id,),
                ).fetchone()
                if exists is None:
                    return {
                        "deleted": False,
                        "conversations_deleted": 0,
                        "documents_deleted": 0,
                        "chunks_deleted": 0,
                    }

                conversations_deleted = int(
                    conn.execute(
                        "SELECT COUNT(*) AS count FROM conversations WHERE folder_id = ?",
                        (folder_id,),
                    ).fetchone()["count"]
                )
                documents_deleted = int(
                    conn.execute(
                        "SELECT COUNT(*) AS count FROM folder_documents WHERE folder_id = ?",
                        (folder_id,),
                    ).fetchone()["count"]
                )
                chunks_deleted = int(
                    conn.execute(
                        """
                        SELECT COUNT(*) AS count
                        FROM folder_chunks c
                        JOIN folder_documents d ON c.document_id = d.id
                        WHERE d.folder_id = ?
                        """,
                        (folder_id,),
                    ).fetchone()["count"]
                )

                conn.execute("DELETE FROM folders WHERE id = ?", (folder_id,))
                conn.commit()
                return {
                    "deleted": True,
                    "conversations_deleted": conversations_deleted,
                    "documents_deleted": documents_deleted,
                    "chunks_deleted": chunks_deleted,
                }

    def _conversation_summary_row(
        self, conn: sqlite3.Connection, conversation_id: str
    ) -> Optional[Dict[str, Any]]:
        row = conn.execute(
            """
            SELECT
                c.id,
                c.title,
                c.folder_id,
                f.name AS folder_name,
                c.created_at,
                c.updated_at,
                c.last_message_at,
                c.last_message_preview,
                c.message_count
            FROM conversations c
            LEFT JOIN folders f ON f.id = c.folder_id
            WHERE c.id = ?
            """,
            (conversation_id,),
        ).fetchone()
        return dict(row) if row else None

    def list_conversations(self) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT
                        c.id,
                        c.title,
                        c.folder_id,
                        f.name AS folder_name,
                        c.created_at,
                        c.updated_at,
                        c.last_message_at,
                        c.last_message_preview,
                        c.message_count
                    FROM conversations c
                    LEFT JOIN folders f ON f.id = c.folder_id
                    ORDER BY c.last_message_at DESC, c.updated_at DESC
                    """
                ).fetchall()
                return [dict(row) for row in rows]

    def create_conversation(self, folder_id: Optional[str]) -> Dict[str, Any]:
        now = self._current_ts()
        conversation_id = str(uuid4())

        with self._lock:
            with self._connect() as conn:
                if folder_id:
                    folder = conn.execute(
                        "SELECT id FROM folders WHERE id = ?",
                        (folder_id,),
                    ).fetchone()
                    if folder is None:
                        raise ValueError("Folder not found")

                conn.execute(
                    """
                    INSERT INTO conversations (
                        id, title, folder_id, created_at, updated_at, last_message_at, last_message_preview, message_count
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        conversation_id,
                        "New chat",
                        folder_id,
                        now,
                        now,
                        now,
                        "",
                        0,
                    ),
                )
                conn.commit()

                summary = self._conversation_summary_row(conn, conversation_id)
                if summary is None:
                    raise ValueError("Failed to create conversation")
                return summary

    def get_conversation(self, conversation_id: str) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                return self._conversation_summary_row(conn, conversation_id)

    def get_conversation_folder_id(self, conversation_id: str) -> Optional[str]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT folder_id FROM conversations WHERE id = ?",
                    (conversation_id,),
                ).fetchone()
                if row is None:
                    return None
                return row["folder_id"]

    def get_conversation_messages(self, conversation_id: str) -> List[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                rows = conn.execute(
                    """
                    SELECT id, conversation_id, role, content, model_used, sources_json, created_at
                    FROM conversation_messages
                    WHERE conversation_id = ?
                    ORDER BY created_at ASC, rowid ASC
                    """,
                    (conversation_id,),
                ).fetchall()

                messages: List[Dict[str, Any]] = []
                for row in rows:
                    item = dict(row)
                    sources_json = item.pop("sources_json")
                    try:
                        item["sources"] = json.loads(sources_json) if sources_json else []
                    except Exception:
                        item["sources"] = []
                    messages.append(item)
                return messages

    def update_conversation(
        self,
        conversation_id: str,
        *,
        title: Optional[str] = None,
        set_folder: bool = False,
        folder_id: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        with self._lock:
            with self._connect() as conn:
                exists = conn.execute(
                    "SELECT id FROM conversations WHERE id = ?",
                    (conversation_id,),
                ).fetchone()
                if exists is None:
                    return None

                assignments: List[str] = []
                params: List[Any] = []

                if title is not None:
                    assignments.append("title = ?")
                    params.append(title.strip())

                if set_folder:
                    if folder_id is not None:
                        folder = conn.execute(
                            "SELECT id FROM folders WHERE id = ?",
                            (folder_id,),
                        ).fetchone()
                        if folder is None:
                            raise ValueError("Folder not found")
                    assignments.append("folder_id = ?")
                    params.append(folder_id)

                if not assignments:
                    return self._conversation_summary_row(conn, conversation_id)

                assignments.append("updated_at = ?")
                params.append(self._current_ts())
                params.append(conversation_id)

                conn.execute(
                    f"UPDATE conversations SET {', '.join(assignments)} WHERE id = ?",
                    tuple(params),
                )
                conn.commit()
                return self._conversation_summary_row(conn, conversation_id)

    def delete_conversation(self, conversation_id: str) -> Dict[str, Any]:
        with self._lock:
            with self._connect() as conn:
                row = conn.execute(
                    "SELECT id FROM conversations WHERE id = ?",
                    (conversation_id,),
                ).fetchone()
                if row is None:
                    return {"deleted": False}

                conn.execute(
                    "DELETE FROM conversations WHERE id = ?",
                    (conversation_id,),
                )
                conn.commit()
                return {"deleted": True}

    def _derive_title_from_message(self, user_message: str) -> str:
        normalized = " ".join(user_message.strip().split())
        if not normalized:
            return "New chat"
        if len(normalized) <= 60:
            return normalized
        return f"{normalized[:57].rstrip()}..."

    def append_turn(
        self,
        conversation_id: str,
        *,
        user_message: str,
        assistant_message: str,
        model_used: Optional[str],
        sources: Optional[List[Dict[str, Any]]],
    ) -> Optional[Dict[str, Any]]:
        now = self._current_ts()

        with self._lock:
            with self._connect() as conn:
                conversation = conn.execute(
                    "SELECT title, message_count FROM conversations WHERE id = ?",
                    (conversation_id,),
                ).fetchone()
                if conversation is None:
                    return None

                conn.execute(
                    """
                    INSERT INTO conversation_messages (
                        id, conversation_id, role, content, model_used, sources_json, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        conversation_id,
                        "user",
                        user_message,
                        None,
                        None,
                        now,
                    ),
                )

                conn.execute(
                    """
                    INSERT INTO conversation_messages (
                        id, conversation_id, role, content, model_used, sources_json, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        str(uuid4()),
                        conversation_id,
                        "assistant",
                        assistant_message,
                        model_used,
                        json.dumps(sources or []),
                        now,
                    ),
                )

                title = str(conversation["title"])
                message_count = int(conversation["message_count"])
                next_title = title
                if title == "New chat" and message_count == 0:
                    next_title = self._derive_title_from_message(user_message)

                preview = " ".join(assistant_message.strip().split())[:200]

                conn.execute(
                    """
                    UPDATE conversations
                    SET
                        title = ?,
                        updated_at = ?,
                        last_message_at = ?,
                        last_message_preview = ?,
                        message_count = message_count + 2
                    WHERE id = ?
                    """,
                    (next_title, now, now, preview, conversation_id),
                )
                conn.commit()
                return self._conversation_summary_row(conn, conversation_id)


PROMPTS_DIR = Path(__file__).parent / "prompts"
BASE_PROMPT = ""
MODE_PROMPTS: Dict[str, str] = {}
PROFILE_PROMPTS: Dict[str, str] = {}
KNOWLEDGE_STORE: Optional[KnowledgeStore] = None
ENV_LOADED = False


class WeatherConfigError(RuntimeError):
    pass


class WeatherAuthError(RuntimeError):
    pass


class WeatherAPIError(RuntimeError):
    pass


def resolve_data_dir() -> Path:
    env_data_dir = os.getenv("VESTA_DATA_DIR")
    if env_data_dir:
        return Path(env_data_dir).expanduser().resolve()
    return (Path.home() / ".vesta").resolve()


def load_vesta_env() -> None:
    global ENV_LOADED
    if ENV_LOADED:
        return

    env_override = os.getenv("VESTA_ENV_FILE")

    # 1) Explicit override path, if provided.
    if env_override:
        override_path = Path(env_override).expanduser().resolve()
        if override_path.exists() and override_path.is_file():
            load_dotenv(override_path, override=False)

    # 2) Data-dir env file.
    data_dir_env = resolve_data_dir() / ".env"
    if data_dir_env.exists() and data_dir_env.is_file():
        load_dotenv(data_dir_env, override=False)

    # 3) Local cwd fallback.
    cwd_env = (Path.cwd() / ".env").resolve()
    if cwd_env.exists() and cwd_env.is_file():
        load_dotenv(cwd_env, override=False)

    ENV_LOADED = True


def get_knowledge_store() -> KnowledgeStore:
    global KNOWLEDGE_STORE
    if KNOWLEDGE_STORE is None:
        KNOWLEDGE_STORE = KnowledgeStore(resolve_data_dir())
    return KNOWLEDGE_STORE


def get_openweather_api_key() -> Optional[str]:
    load_vesta_env()
    api_key = os.getenv("OPENWEATHER_API_KEY", "").strip()
    return api_key or None


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def day_key_from_epoch(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).date().isoformat()


def parse_int_timestamp(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(str(value))
    except Exception:
        return None


def is_weather_intent(message: str) -> bool:
    if not message.strip():
        return False
    return WEATHER_INTENT_PATTERN.search(message) is not None


def clamp_score(value: float, min_value: float = 0.0, max_value: float = 100.0) -> float:
    return max(min_value, min(max_value, value))


def normalize_weather_mode(mode: Optional[str]) -> str:
    normalized = (mode or DEFAULT_WEATHER_MODE).strip().lower()
    return normalized if normalized in WEATHER_MODE_OPTIONS else DEFAULT_WEATHER_MODE


def normalize_country_code(country: Optional[str]) -> str:
    raw = (country or "US").strip().upper()
    if not raw:
        return "US"
    return raw[:3]


def make_weather_bucket(temp_f: float, wind_mph: float, precip_in: float) -> str:
    if temp_f < 45:
        temp_bin = "cold"
    elif temp_f > 85:
        temp_bin = "hot"
    else:
        temp_bin = "mild"

    if wind_mph < 10:
        wind_bin = "calm"
    elif wind_mph < 20:
        wind_bin = "breezy"
    else:
        wind_bin = "windy"

    precip_bin = "wet" if precip_in >= 0.1 else "dry"
    return f"{temp_bin}:{wind_bin}:{precip_bin}"


def normalize_current_weather(payload: Dict[str, Any]) -> Dict[str, Any]:
    main = payload.get("main") if isinstance(payload.get("main"), dict) else {}
    wind = payload.get("wind") if isinstance(payload.get("wind"), dict) else {}
    rain = payload.get("rain") if isinstance(payload.get("rain"), dict) else {}
    snow = payload.get("snow") if isinstance(payload.get("snow"), dict) else {}
    weather_items = payload.get("weather") if isinstance(payload.get("weather"), list) else []
    weather_head = weather_items[0] if weather_items and isinstance(weather_items[0], dict) else {}

    precip = 0.0
    if rain:
        precip += float(rain.get("1h") or 0.0)
    if snow:
        precip += float(snow.get("1h") or 0.0)

    return {
        "observed_ts": int(payload.get("dt") or time.time()),
        "temp_f": float(main.get("temp")) if main.get("temp") is not None else None,
        "feels_like_f": float(main.get("feels_like"))
        if main.get("feels_like") is not None
        else None,
        "humidity_pct": float(main.get("humidity"))
        if main.get("humidity") is not None
        else None,
        "wind_mph": float(wind.get("speed")) if wind.get("speed") is not None else None,
        "wind_gust_mph": float(wind.get("gust")) if wind.get("gust") is not None else None,
        "precip_in": precip,
        "condition_code": int(weather_head.get("id")) if weather_head.get("id") is not None else None,
        "condition_main": weather_head.get("main"),
        "condition_desc": weather_head.get("description"),
        "raw_json": payload,
    }


def normalize_forecast_points(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    items = payload.get("list") if isinstance(payload.get("list"), list) else []
    points: List[Dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue

        main = item.get("main") if isinstance(item.get("main"), dict) else {}
        wind = item.get("wind") if isinstance(item.get("wind"), dict) else {}
        rain = item.get("rain") if isinstance(item.get("rain"), dict) else {}
        snow = item.get("snow") if isinstance(item.get("snow"), dict) else {}
        weather_items = item.get("weather") if isinstance(item.get("weather"), list) else []
        weather_head = weather_items[0] if weather_items and isinstance(weather_items[0], dict) else {}

        precip = 0.0
        if rain:
            precip += float(rain.get("3h") or rain.get("1h") or 0.0)
        if snow:
            precip += float(snow.get("3h") or snow.get("1h") or 0.0)

        forecast_ts = int(item.get("dt") or time.time())
        hours_out = max((forecast_ts - int(time.time())) / 3600.0, 0.0)
        horizon_decay = max(0.0, 1.0 - (hours_out / 120.0))
        confidence = clamp_score(horizon_decay * 100.0) / 100.0

        points.append(
            {
                "forecast_ts": str(forecast_ts),
                "temp_f": float(main.get("temp")) if main.get("temp") is not None else None,
                "temp_min_f": float(main.get("temp_min"))
                if main.get("temp_min") is not None
                else None,
                "temp_max_f": float(main.get("temp_max"))
                if main.get("temp_max") is not None
                else None,
                "humidity_pct": float(main.get("humidity"))
                if main.get("humidity") is not None
                else None,
                "wind_mph": float(wind.get("speed")) if wind.get("speed") is not None else None,
                "precip_prob": float(item.get("pop")) if item.get("pop") is not None else 0.0,
                "precip_in": precip,
                "condition_code": int(weather_head.get("id"))
                if weather_head.get("id") is not None
                else None,
                "condition_main": weather_head.get("main"),
                "condition_desc": weather_head.get("description"),
                "confidence_score": confidence,
                "raw_json": item,
            }
        )

    return points


def normalize_alerts(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    alerts = payload.get("alerts")
    if not isinstance(alerts, list):
        return []

    normalized: List[Dict[str, Any]] = []
    for item in alerts:
        if not isinstance(item, dict):
            continue
        normalized.append(
            {
                "event": item.get("event"),
                "severity": item.get("severity"),
                "sender_name": item.get("sender_name"),
                "start_ts": str(item.get("start")) if item.get("start") is not None else None,
                "end_ts": str(item.get("end")) if item.get("end") is not None else None,
                "description": item.get("description"),
                "raw_json": item,
            }
        )
    return normalized


def summarize_daily_forecast(points: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for point in points:
        ts = parse_int_timestamp(point.get("forecast_ts"))
        if ts is None:
            continue
        grouped[day_key_from_epoch(ts)].append(point)

    daily: List[Dict[str, Any]] = []
    for date_key in sorted(grouped.keys())[:5]:
        rows = grouped[date_key]
        temps = [float(row["temp_f"]) for row in rows if row.get("temp_f") is not None]
        winds = [float(row["wind_mph"]) for row in rows if row.get("wind_mph") is not None]
        pops = [float(row.get("precip_prob") or 0.0) for row in rows]
        precips = [float(row.get("precip_in") or 0.0) for row in rows]
        confidences = [float(row.get("confidence_score") or 0.0) for row in rows]
        first = rows[0]

        daily.append(
            {
                "date": date_key,
                "temp_min_f": min(temps) if temps else None,
                "temp_max_f": max(temps) if temps else None,
                "temp_avg_f": sum(temps) / len(temps) if temps else None,
                "wind_max_mph": max(winds) if winds else None,
                "precip_total_in": sum(precips),
                "precip_prob_avg": sum(pops) / len(pops) if pops else 0.0,
                "confidence_score": sum(confidences) / len(confidences) if confidences else 0.0,
                "condition_main": first.get("condition_main"),
                "condition_desc": first.get("condition_desc"),
            }
        )
    return daily


def hydrate_forecast_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    points: List[Dict[str, Any]] = []
    for row in rows:
        points.append(
            {
                "forecast_ts": row.get("forecast_ts"),
                "temp_f": row.get("temp_f"),
                "temp_min_f": row.get("temp_min_f"),
                "temp_max_f": row.get("temp_max_f"),
                "humidity_pct": row.get("humidity_pct"),
                "wind_mph": row.get("wind_mph"),
                "precip_prob": row.get("precip_prob"),
                "precip_in": row.get("precip_in"),
                "condition_code": row.get("condition_code"),
                "condition_main": row.get("condition_main"),
                "condition_desc": row.get("condition_desc"),
                "confidence_score": row.get("confidence_score"),
            }
        )
    return points


def cache_age_seconds(last_refresh_ts: Optional[str]) -> Optional[int]:
    latest = parse_int_timestamp(last_refresh_ts)
    if latest is None:
        return None
    age = int(time.time()) - latest
    return max(age, 0)


def is_cache_stale(last_refresh_ts: Optional[str], ttl_minutes: int) -> bool:
    age = cache_age_seconds(last_refresh_ts)
    if age is None:
        return True
    return age > int(ttl_minutes * 60)


async def call_openweather_json(
    *,
    url: str,
    endpoint_name: str,
    params: Dict[str, Any],
    include_units: bool = True,
    expect_list: bool = False,
) -> Any:
    api_key = get_openweather_api_key()
    if not api_key:
        raise WeatherConfigError("OpenWeather API key is not configured.")

    request_params = {key: value for key, value in params.items() if value is not None}
    request_params["appid"] = api_key
    if include_units and "units" not in request_params:
        request_params["units"] = "imperial"

    store = get_knowledge_store()
    start = time.time()
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            response = await client.get(url, params=request_params)
        latency_ms = (time.time() - start) * 1000

        if response.status_code == 401:
            error_text = response.text[:500]
            store.log_weather_api_call(
                endpoint=endpoint_name,
                success=False,
                http_status=401,
                latency_ms=latency_ms,
                error_text=error_text,
            )

            lowered = error_text.lower()
            if endpoint_name == "alerts" and (
                "one call 3.0 requires a separate subscription" in lowered
                or "one call by call plan" in lowered
            ):
                raise WeatherAPIError(
                    "OpenWeather One Call 3.0 subscription is required for alerts."
                )

            raise WeatherAuthError("OpenWeather API key is invalid.")

        if response.status_code >= 400:
            store.log_weather_api_call(
                endpoint=endpoint_name,
                success=False,
                http_status=response.status_code,
                latency_ms=latency_ms,
                error_text=response.text[:500],
            )
            raise WeatherAPIError(
                f"OpenWeather request failed for {endpoint_name}: status {response.status_code}"
            )

        payload = response.json()
        store.log_weather_api_call(
            endpoint=endpoint_name,
            success=True,
            http_status=response.status_code,
            latency_ms=latency_ms,
            error_text=None,
        )
        if expect_list:
            if not isinstance(payload, list):
                raise WeatherAPIError("OpenWeather payload was not a JSON array.")
        elif not isinstance(payload, dict):
            raise WeatherAPIError("OpenWeather payload was not a JSON object.")
        return payload
    except WeatherAuthError:
        raise
    except WeatherConfigError:
        raise
    except WeatherAPIError:
        raise
    except Exception as error:
        latency_ms = (time.time() - start) * 1000
        store.log_weather_api_call(
            endpoint=endpoint_name,
            success=False,
            http_status=None,
            latency_ms=latency_ms,
            error_text=str(error),
        )
        raise WeatherAPIError(str(error))


async def fetch_openweather_current(lat: float, lon: float) -> Dict[str, Any]:
    payload = await call_openweather_json(
        url=OPENWEATHER_CURRENT_URL,
        endpoint_name="current",
        params={"lat": lat, "lon": lon},
    )
    return normalize_current_weather(payload)


async def fetch_openweather_forecast(lat: float, lon: float) -> List[Dict[str, Any]]:
    payload = await call_openweather_json(
        url=OPENWEATHER_FORECAST_URL,
        endpoint_name="forecast",
        params={"lat": lat, "lon": lon},
    )
    return normalize_forecast_points(payload)


async def fetch_openweather_alerts(lat: float, lon: float) -> List[Dict[str, Any]]:
    payload = await call_openweather_json(
        url=OPENWEATHER_ONECALL_URL,
        endpoint_name="alerts",
        params={"lat": lat, "lon": lon, "exclude": "minutely,hourly,daily,current"},
    )
    return normalize_alerts(payload)


async def resolve_openweather_location(
    *,
    city: str,
    state: Optional[str],
    country: str,
) -> List[Dict[str, Any]]:
    query_parts = [city.strip()]
    if state and state.strip():
        query_parts.append(state.strip())
    if country.strip():
        query_parts.append(country.strip().upper())
    query = ",".join(query_parts)

    payload = await call_openweather_json(
        url=OPENWEATHER_GEOCODE_URL,
        endpoint_name="geocode",
        params={"q": query, "limit": 5},
        include_units=False,
        expect_list=True,
    )
    if not isinstance(payload, list):
        return []

    results: List[Dict[str, Any]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        if item.get("lat") is None or item.get("lon") is None:
            continue
        results.append(
            {
                "name": str(item.get("name") or city).strip(),
                "state": str(item.get("state") or "").strip() or None,
                "country": str(item.get("country") or country).strip().upper(),
                "lat": float(item["lat"]),
                "lon": float(item["lon"]),
            }
        )
    return results


def normalize_current_row(row: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not row:
        return None
    return {
        "observed_ts": parse_int_timestamp(row.get("fetched_at")) or int(time.time()),
        "temp_f": row.get("temp_f"),
        "feels_like_f": row.get("feels_like_f"),
        "humidity_pct": row.get("humidity_pct"),
        "wind_mph": row.get("wind_mph"),
        "wind_gust_mph": row.get("wind_gust_mph"),
        "precip_in": row.get("precip_in") or 0.0,
        "condition_code": row.get("condition_code"),
        "condition_main": row.get("condition_main"),
        "condition_desc": row.get("condition_desc"),
    }


def normalize_alert_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    alerts: List[Dict[str, Any]] = []
    for row in rows:
        alerts.append(
            {
                "event": row.get("event"),
                "severity": row.get("severity"),
                "sender_name": row.get("sender_name"),
                "start_ts": row.get("start_ts"),
                "end_ts": row.get("end_ts"),
                "description": row.get("description"),
            }
        )
    return alerts


def latest_weather_fetched_ts(
    current_row: Optional[Dict[str, Any]],
    forecast_rows: List[Dict[str, Any]],
    alert_rows: List[Dict[str, Any]],
) -> Optional[str]:
    candidates: List[str] = []
    if current_row and current_row.get("fetched_at"):
        candidates.append(str(current_row["fetched_at"]))
    if forecast_rows and forecast_rows[0].get("fetched_at"):
        candidates.append(str(forecast_rows[0]["fetched_at"]))
    if alert_rows and alert_rows[0].get("fetched_at"):
        candidates.append(str(alert_rows[0]["fetched_at"]))
    if not candidates:
        return None
    return max(candidates)


def get_cached_weather_bundle(lat: float, lon: float) -> Dict[str, Any]:
    store = get_knowledge_store()
    current_row = store.get_latest_weather_current(lat, lon)
    forecast_rows = store.get_latest_weather_forecast(lat, lon)
    alert_rows = store.get_latest_weather_alerts(lat, lon)
    fetched_at = latest_weather_fetched_ts(current_row, forecast_rows, alert_rows)

    return {
        "current": normalize_current_row(current_row),
        "forecast_3h": hydrate_forecast_rows(forecast_rows),
        "forecast_daily": summarize_daily_forecast(hydrate_forecast_rows(forecast_rows)),
        "alerts": normalize_alert_rows(alert_rows),
        "last_refresh_ts": fetched_at,
    }


def calculate_data_completeness(
    current: Optional[Dict[str, Any]],
    forecast_points: List[Dict[str, Any]],
) -> float:
    required_current = [
        current.get("temp_f") if current else None,
        current.get("humidity_pct") if current else None,
        current.get("wind_mph") if current else None,
        current.get("condition_code") if current else None,
    ]
    current_score = sum(1 for value in required_current if value is not None) / max(
        len(required_current), 1
    )
    forecast_score = 1.0 if forecast_points else 0.0
    return max(0.0, min(1.0, (current_score * 0.7) + (forecast_score * 0.3)))


def calculate_forecast_confidence(forecast_points: List[Dict[str, Any]]) -> float:
    if not forecast_points:
        return 0.0

    now_epoch = int(time.time())
    values: List[float] = []
    for point in forecast_points:
        ts = parse_int_timestamp(point.get("forecast_ts"))
        if ts is None:
            continue
        hours_out = max((ts - now_epoch) / 3600.0, 0.0)
        horizon_decay = max(0.0, 1.0 - (hours_out / 120.0))
        confidence = point.get("confidence_score")
        if confidence is None:
            confidence = horizon_decay
        values.append(max(0.0, min(1.0, float(confidence))))

    if not values:
        return 0.0
    return sum(values) / len(values)


def calculate_integrity_score(
    *,
    forecast_points: List[Dict[str, Any]],
    current: Optional[Dict[str, Any]],
    api_reliability: float,
) -> float:
    forecast_confidence = calculate_forecast_confidence(forecast_points)
    completeness = calculate_data_completeness(current, forecast_points)
    integrity = 100.0 * (
        (forecast_confidence * 0.4) + (completeness * 0.3) + (api_reliability * 0.3)
    )
    return round(clamp_score(integrity), 2)


def calculate_mode_meaning_score(
    *,
    mode: str,
    current: Optional[Dict[str, Any]],
    forecast_daily: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]],
) -> float:
    mode_key = normalize_weather_mode(mode)
    first_day = forecast_daily[0] if forecast_daily else {}
    temp = float(first_day.get("temp_avg_f") or current.get("temp_f") or 70.0) if current else float(first_day.get("temp_avg_f") or 70.0)
    wind = float(first_day.get("wind_max_mph") or current.get("wind_mph") or 0.0) if current else float(first_day.get("wind_max_mph") or 0.0)
    precip = float(first_day.get("precip_total_in") or current.get("precip_in") or 0.0) if current else float(first_day.get("precip_total_in") or 0.0)

    alert_text = " ".join(
        f"{item.get('event', '')} {item.get('description', '')}".lower() for item in alerts
    )
    hail_signal = 1.0 if ("hail" in alert_text or "thunderstorm" in alert_text) else 0.0

    if mode_key == "storm_damage":
        wind_score = max(0.0, min(1.0, wind / 45.0))
        precip_score = max(0.0, min(1.0, precip / 1.0))
        score = (wind_score * 0.4) + (precip_score * 0.3) + (hail_signal * 0.3)
        return round(clamp_score(score * 100.0), 2)

    if mode_key == "lawn_care":
        precip_balance = max(0.0, 1.0 - min(abs(precip - 0.2) / 0.25, 1.0))
        moisture_estimate = max(0.0, min(1.0, precip / 0.6))
        temp_ideal = max(0.0, 1.0 - min(abs(temp - 72.0) / 25.0, 1.0))
        score = (precip_balance * 0.35) + (moisture_estimate * 0.25) + (temp_ideal * 0.4)
        return round(clamp_score(score * 100.0), 2)

    if mode_key == "construction":
        precip_disrupt = max(0.0, min(1.0, precip / 0.35))
        wind_disrupt = max(0.0, min(1.0, wind / 28.0))
        temp_disrupt = max(0.0, min(1.0, abs(temp - 68.0) / 35.0))
        score = (precip_disrupt * 0.4) + (wind_disrupt * 0.3) + (temp_disrupt * 0.3)
        return round(clamp_score(score * 100.0), 2)

    # general
    precip_signal = max(0.0, min(1.0, precip / 0.4))
    wind_signal = max(0.0, min(1.0, wind / 30.0))
    temp_signal = max(0.0, min(1.0, abs(temp - 70.0) / 30.0))
    score = (precip_signal * 0.35) + (wind_signal * 0.3) + (temp_signal * 0.35)
    return round(clamp_score(score * 100.0), 2)


def calculate_resilience_score(
    *,
    mode: str,
    forecast_daily: List[Dict[str, Any]],
) -> float:
    store = get_knowledge_store()
    history = store.list_prediction_accuracy(normalize_weather_mode(mode))
    if not history:
        return 65.0

    day = forecast_daily[0] if forecast_daily else {}
    temp = float(day.get("temp_avg_f") or 70.0)
    wind = float(day.get("wind_max_mph") or 0.0)
    precip = float(day.get("precip_total_in") or 0.0)
    target_bucket = make_weather_bucket(temp, wind, precip)

    bucket_scores: List[float] = []
    overall_scores: List[float] = []
    for item in history:
        score = item.get("accuracy_score")
        if score is None:
            continue
        score_value = float(score)
        overall_scores.append(score_value)
        factors = item.get("factors") if isinstance(item.get("factors"), dict) else {}
        if factors.get("bucket") == target_bucket:
            bucket_scores.append(score_value)

    if bucket_scores:
        return round(clamp_score((sum(bucket_scores) / len(bucket_scores)) * 100.0), 2)
    if overall_scores:
        return round(clamp_score((sum(overall_scores) / len(overall_scores)) * 100.0), 2)
    return 65.0


def calculate_cci(
    *,
    mode: str,
    integrity: float,
    resilience: float,
    meaning: float,
) -> float:
    mode_key = normalize_weather_mode(mode)
    if mode_key == "storm_damage":
        value = (integrity * 0.25) + (resilience * 0.30) + (meaning * 0.45)
    elif mode_key == "lawn_care":
        value = (integrity * 0.30) + (resilience * 0.30) + (meaning * 0.40)
    elif mode_key == "construction":
        value = (integrity * 0.30) + (resilience * 0.35) + (meaning * 0.35)
    else:
        value = (integrity * 0.34) + (resilience * 0.33) + (meaning * 0.33)
    return round(clamp_score(value), 2)


def calculate_coherence_scores(
    *,
    mode: str,
    current: Optional[Dict[str, Any]],
    forecast_points: List[Dict[str, Any]],
    forecast_daily: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]],
) -> Dict[str, float]:
    store = get_knowledge_store()
    api_reliability = store.get_weather_api_reliability()
    integrity = calculate_integrity_score(
        forecast_points=forecast_points,
        current=current,
        api_reliability=api_reliability,
    )
    resilience = calculate_resilience_score(mode=mode, forecast_daily=forecast_daily)
    meaning = calculate_mode_meaning_score(
        mode=mode,
        current=current,
        forecast_daily=forecast_daily,
        alerts=alerts,
    )
    cci = calculate_cci(mode=mode, integrity=integrity, resilience=resilience, meaning=meaning)
    return {
        "integrity": integrity,
        "resilience": resilience,
        "meaning": meaning,
        "cci": cci,
    }


def generate_mode_prediction(
    *,
    mode: str,
    day: Dict[str, Any],
    day_offset: int,
) -> Dict[str, Any]:
    temp = float(day.get("temp_avg_f") or 70.0)
    wind = float(day.get("wind_max_mph") or 0.0)
    precip = float(day.get("precip_total_in") or 0.0)
    confidence = float(day.get("confidence_score") or 0.6)
    mode_key = normalize_weather_mode(mode)

    if mode_key == "storm_damage":
        probability = min(100.0, ((wind / 45.0) * 45.0) + ((precip / 1.0) * 35.0) + 10.0)
        text = f"{probability:.0f}% probability of storm damage opportunity in {day_offset} day(s)"
    elif mode_key == "lawn_care":
        temp_fit = max(0.0, 1.0 - min(abs(temp - 72.0) / 22.0, 1.0))
        rain_fit = max(0.0, 1.0 - min(abs(precip - 0.2) / 0.3, 1.0))
        probability = min(100.0, (temp_fit * 60.0) + (rain_fit * 40.0))
        text = f"{probability:.0f}% probability of optimal lawn care window in {day_offset} day(s)"
    elif mode_key == "construction":
        disruption = min(1.0, (precip / 0.35) * 0.45 + (wind / 28.0) * 0.35 + (abs(temp - 68.0) / 35.0) * 0.2)
        probability = disruption * 100.0
        text = f"{probability:.0f}% probability of construction delay in {day_offset} day(s)"
    else:
        shift = min(1.0, (abs(temp - 70.0) / 30.0) * 0.4 + (precip / 0.4) * 0.35 + (wind / 30.0) * 0.25)
        probability = shift * 100.0
        text = f"{probability:.0f}% probability of notable weather shift in {day_offset} day(s)"

    probability = round(clamp_score(probability * confidence), 2)
    bucket = make_weather_bucket(temp, wind, precip)
    return {
        "target_date": day["date"],
        "probability": probability,
        "prediction_text": text,
        "factors": {
            "bucket": bucket,
            "temp_f": temp,
            "wind_mph": wind,
            "precip_in": precip,
            "confidence": confidence,
        },
    }


def generate_predictions(
    *,
    mode: str,
    forecast_daily: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    for index, day in enumerate(forecast_daily[:5]):
        results.append(
            generate_mode_prediction(
                mode=mode,
                day=day,
                day_offset=index + 1,
            )
        )
    return results


def did_prediction_outcome_happen(mode: str, factors: Dict[str, Any]) -> bool:
    mode_key = normalize_weather_mode(mode)
    temp = float(factors.get("temp_f") or 70.0)
    wind = float(factors.get("wind_mph") or 0.0)
    precip = float(factors.get("precip_in") or 0.0)

    if mode_key == "storm_damage":
        return wind >= 35.0 or precip >= 0.75
    if mode_key == "lawn_care":
        return 60.0 <= temp <= 82.0 and 0.05 <= precip <= 0.45
    if mode_key == "construction":
        return precip >= 0.25 or wind >= 22.0 or temp <= 35.0 or temp >= 95.0
    return precip >= 0.3 or wind >= 20.0 or abs(temp - 70.0) >= 18.0


def evaluate_mature_predictions() -> int:
    store = get_knowledge_store()
    today = now_utc().date().isoformat()
    pending = store.list_mature_predictions_without_outcome(today)
    written = 0
    for prediction in pending:
        probability = float(prediction.get("probability") or 0.0)
        actual_happened = did_prediction_outcome_happen(
            str(prediction.get("mode") or DEFAULT_WEATHER_MODE),
            prediction.get("factors") if isinstance(prediction.get("factors"), dict) else {},
        )
        if actual_happened:
            accuracy = max(0.0, min(1.0, probability / 100.0))
        else:
            accuracy = max(0.0, min(1.0, (100.0 - probability) / 100.0))
        store.save_prediction_outcome(
            prediction_id=str(prediction["id"]),
            actual_outcome=actual_happened,
            accuracy_score=round(accuracy, 4),
            notes="Auto-evaluated from forecast factors heuristic",
        )
        written += 1
    return written


def build_weather_insights(
    *,
    mode: str,
    coherence: Dict[str, float],
    current: Optional[Dict[str, Any]],
    forecast_daily: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]],
) -> List[str]:
    mode_key = normalize_weather_mode(mode)
    insights: List[str] = []

    if coherence["cci"] >= 75:
        insights.append("Coherence is high: forecast confidence and mode alignment are strong.")
    elif coherence["cci"] >= 55:
        insights.append("Coherence is moderate: conditions are usable with some uncertainty.")
    else:
        insights.append("Coherence is low: treat near-term predictions as tentative.")

    if alerts:
        insights.append(f"{len(alerts)} weather alert(s) are active for this location.")

    if current:
        temp = current.get("temp_f")
        wind = current.get("wind_mph")
        precip = current.get("precip_in")
        if temp is not None and wind is not None:
            insights.append(
                f"Current conditions: {temp:.0f}F, wind {wind:.0f} mph, precip {float(precip or 0.0):.2f} in."
            )

    if forecast_daily:
        first = forecast_daily[0]
        insights.append(
            f"Next-day outlook: high {float(first.get('temp_max_f') or 0.0):.0f}F, "
            f"precip {float(first.get('precip_total_in') or 0.0):.2f} in."
        )

    if mode_key == "storm_damage":
        insights.append("Storm mode prioritizes wind, precip intensity, and severe-signal markers.")
    elif mode_key == "lawn_care":
        insights.append("Lawn mode favors moderate moisture and ideal turf temperature windows.")
    elif mode_key == "construction":
        insights.append("Construction mode emphasizes delay risk from rain, wind, and extreme temperatures.")
    else:
        insights.append("General mode balances overall weather significance across core signals.")

    return insights


async def refresh_weather_bundle(
    *,
    lat: float,
    lon: float,
    force_refresh: bool,
    ttl_minutes: int,
) -> Dict[str, Any]:
    cache_bundle = get_cached_weather_bundle(lat, lon)
    cache_ts = cache_bundle["last_refresh_ts"]
    has_cache = cache_bundle["current"] is not None or bool(cache_bundle["forecast_3h"])
    stale = is_cache_stale(cache_ts, ttl_minutes)

    if not force_refresh and has_cache and not stale:
        return {
            **cache_bundle,
            "source": "cache",
            "stale": False,
            "warning": None,
        }

    if not get_openweather_api_key():
        if has_cache:
            return {
                **cache_bundle,
                "source": "cache",
                "stale": True,
                "warning": "missing_api_key",
            }
        raise WeatherConfigError("OpenWeather API key is not configured.")

    try:
        fetched_at = str(int(time.time()))
        current = await fetch_openweather_current(lat, lon)
        forecast_points = await fetch_openweather_forecast(lat, lon)
        alerts: List[Dict[str, Any]] = []
        warning: Optional[str] = None
        try:
            alerts = await fetch_openweather_alerts(lat, lon)
        except Exception as alert_error:
            warning = f"alerts_unavailable: {alert_error}"
            print(
                f"Weather alerts warning: {type(alert_error).__name__}: {alert_error}",
                file=sys.stderr,
            )
        store = get_knowledge_store()
        store.save_weather_snapshot(
            lat=lat,
            lon=lon,
            current=current,
            forecast_points=forecast_points,
            alerts=alerts,
            fetched_at=fetched_at,
        )
        return {
            "current": current,
            "forecast_3h": forecast_points,
            "forecast_daily": summarize_daily_forecast(forecast_points),
            "alerts": alerts,
            "last_refresh_ts": fetched_at,
            "source": "live",
            "stale": False,
            "warning": warning,
        }
    except Exception as error:
        if has_cache:
            return {
                **cache_bundle,
                "source": "cache",
                "stale": True,
                "warning": str(error),
            }
        raise


def get_weather_status_payload() -> Dict[str, Any]:
    store = get_knowledge_store()
    has_cache = store.has_cached_weather_data()
    last_refresh = store.get_last_weather_refresh_ts()
    api_key = get_openweather_api_key()
    if not api_key:
        return {
            "enabled": False,
            "reason": "missing_api_key",
            "has_cached_data": has_cache,
            "last_refresh_ts": last_refresh,
        }

    latest_event = store.get_latest_weather_api_event()
    reason = None
    if latest_event and not bool(latest_event.get("success")):
        status = latest_event.get("http_status")
        error_text = str(latest_event.get("error_text") or "").lower()
        if status == 401:
            if "one call" in error_text and "subscription" in error_text:
                reason = None
            else:
                reason = "invalid_api_key"
        elif status is None or int(status) >= 500:
            reason = "api_unreachable"

    return {
        "enabled": reason is None,
        "reason": reason,
        "has_cached_data": has_cache,
        "last_refresh_ts": last_refresh,
    }


def build_weather_context_block(
    *,
    mode: str,
    location: Dict[str, Any],
    current: Optional[Dict[str, Any]],
    forecast_daily: List[Dict[str, Any]],
    alerts: List[Dict[str, Any]],
    coherence: Dict[str, float],
    predictions: List[Dict[str, Any]],
) -> Tuple[str, List[Dict[str, Any]]]:
    if not current and not forecast_daily and not predictions:
        return "", []

    location_label = (
        f"{location.get('city')}, {location.get('state')}" if location.get("state") else location.get("city")
    )
    context_parts = [
        "Weather Intelligence Context:",
        f"Location: {location_label}, {location.get('country')}",
        f"Mode: {mode}",
        f"Coherence: integrity={coherence['integrity']:.1f}, resilience={coherence['resilience']:.1f}, "
        f"meaning={coherence['meaning']:.1f}, cci={coherence['cci']:.1f}",
        "",
    ]
    sources: List[Dict[str, Any]] = []

    if current:
        context_parts.append(
            f"Current: {float(current.get('temp_f') or 0.0):.0f}F, "
            f"wind {float(current.get('wind_mph') or 0.0):.0f} mph, "
            f"precip {float(current.get('precip_in') or 0.0):.2f} in, "
            f"{current.get('condition_desc') or current.get('condition_main') or 'unknown'}."
        )
        sources.append(
            {
                "source_type": "weather",
                "label": "Current conditions",
                "observed_at": str(current.get("observed_ts") or int(time.time())),
                "mode": mode,
            }
        )

    for index, day in enumerate(forecast_daily[:3]):
        context_parts.append(
            f"Day {index + 1} ({day['date']}): high {float(day.get('temp_max_f') or 0.0):.0f}F, "
            f"precip {float(day.get('precip_total_in') or 0.0):.2f} in, "
            f"wind max {float(day.get('wind_max_mph') or 0.0):.0f} mph."
        )
        sources.append(
            {
                "source_type": "weather",
                "label": f"Forecast day {index + 1}",
                "observed_at": day["date"],
                "mode": mode,
            }
        )

    if alerts:
        context_parts.append(f"Alerts: {len(alerts)} active weather alert(s).")
        sources.append(
            {
                "source_type": "weather",
                "label": "Severe weather alerts",
                "observed_at": str(int(time.time())),
                "mode": mode,
            }
        )

    for prediction in predictions[:2]:
        context_parts.append(f"Prediction: {prediction['prediction_text']}")
        sources.append(
            {
                "source_type": "weather",
                "label": "Prediction",
                "observed_at": prediction.get("target_date"),
                "mode": mode,
            }
        )

    return "\n".join(context_parts).strip(), sources


async def retrieve_weather_context(query: str) -> Tuple[str, List[Dict[str, Any]]]:
    if not is_weather_intent(query):
        return "", []

    settings = get_knowledge_store().get_weather_settings()
    location = settings.get("location")
    if not location:
        return "", []

    mode = normalize_weather_mode(settings.get("mode"))
    bundle = await refresh_weather_bundle(
        lat=float(location["lat"]),
        lon=float(location["lon"]),
        force_refresh=False,
        ttl_minutes=int(settings.get("cache_ttl_minutes") or WEATHER_CACHE_TTL_MINUTES),
    )
    coherence = calculate_coherence_scores(
        mode=mode,
        current=bundle.get("current"),
        forecast_points=bundle.get("forecast_3h") or [],
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
    )
    predictions = generate_predictions(
        mode=mode,
        forecast_daily=bundle.get("forecast_daily") or [],
    )

    return build_weather_context_block(
        mode=mode,
        location=location,
        current=bundle.get("current"),
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
        coherence=coherence,
        predictions=predictions,
    )


async def get_or_refresh_current_weather(
    *,
    lat: float,
    lon: float,
    force_refresh: bool,
    ttl_minutes: int,
) -> Dict[str, Any]:
    store = get_knowledge_store()
    current_row = store.get_latest_weather_current(lat, lon)
    current = normalize_current_row(current_row)
    cache_ts = str(current_row["fetched_at"]) if current_row and current_row.get("fetched_at") else None
    stale = is_cache_stale(cache_ts, ttl_minutes)

    if current and not force_refresh and not stale:
        return {"data": current, "source": "cache", "stale": False, "fetched_at": cache_ts}

    if not get_openweather_api_key():
        if current:
            return {"data": current, "source": "cache", "stale": True, "fetched_at": cache_ts}
        raise WeatherConfigError("OpenWeather API key is not configured.")

    try:
        live = await fetch_openweather_current(lat, lon)
        fetched_at = store.save_current_weather(lat=lat, lon=lon, current=live)
        return {"data": live, "source": "live", "stale": False, "fetched_at": fetched_at}
    except Exception:
        if current:
            return {"data": current, "source": "cache", "stale": True, "fetched_at": cache_ts}
        raise


async def get_or_refresh_forecast_weather(
    *,
    lat: float,
    lon: float,
    force_refresh: bool,
    ttl_minutes: int,
) -> Dict[str, Any]:
    store = get_knowledge_store()
    rows = store.get_latest_weather_forecast(lat, lon)
    points = hydrate_forecast_rows(rows)
    cache_ts = str(rows[0]["fetched_at"]) if rows and rows[0].get("fetched_at") else None
    stale = is_cache_stale(cache_ts, ttl_minutes)

    if points and not force_refresh and not stale:
        return {
            "forecast_3h": points,
            "forecast_daily": summarize_daily_forecast(points),
            "source": "cache",
            "stale": False,
            "fetched_at": cache_ts,
        }

    if not get_openweather_api_key():
        if points:
            return {
                "forecast_3h": points,
                "forecast_daily": summarize_daily_forecast(points),
                "source": "cache",
                "stale": True,
                "fetched_at": cache_ts,
            }
        raise WeatherConfigError("OpenWeather API key is not configured.")

    try:
        live_points = await fetch_openweather_forecast(lat, lon)
        fetched_at = store.save_forecast_weather(
            lat=lat,
            lon=lon,
            forecast_points=live_points,
        )
        return {
            "forecast_3h": live_points,
            "forecast_daily": summarize_daily_forecast(live_points),
            "source": "live",
            "stale": False,
            "fetched_at": fetched_at,
        }
    except Exception:
        if points:
            return {
                "forecast_3h": points,
                "forecast_daily": summarize_daily_forecast(points),
                "source": "cache",
                "stale": True,
                "fetched_at": cache_ts,
            }
        raise


async def get_or_refresh_alert_weather(
    *,
    lat: float,
    lon: float,
    force_refresh: bool,
    ttl_minutes: int,
) -> Dict[str, Any]:
    store = get_knowledge_store()
    rows = store.get_latest_weather_alerts(lat, lon)
    alerts = normalize_alert_rows(rows)
    cache_ts = str(rows[0]["fetched_at"]) if rows and rows[0].get("fetched_at") else None
    stale = is_cache_stale(cache_ts, ttl_minutes)

    if alerts and not force_refresh and not stale:
        return {"alerts": alerts, "source": "cache", "stale": False, "fetched_at": cache_ts}

    if not get_openweather_api_key():
        if alerts:
            return {"alerts": alerts, "source": "cache", "stale": True, "fetched_at": cache_ts}
        raise WeatherConfigError("OpenWeather API key is not configured.")

    try:
        live_alerts = await fetch_openweather_alerts(lat, lon)
        fetched_at = store.save_weather_alerts(lat=lat, lon=lon, alerts=live_alerts)
        return {"alerts": live_alerts, "source": "live", "stale": False, "fetched_at": fetched_at}
    except Exception:
        if alerts:
            return {"alerts": alerts, "source": "cache", "stale": True, "fetched_at": cache_ts}
        raise


def chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    cleaned = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if not cleaned:
        return []

    chunks: List[str] = []
    step = max(size - overlap, 1)
    start = 0
    while start < len(cleaned):
        chunk = cleaned[start : start + size].strip()
        if chunk:
            chunks.append(chunk)
        if start + size >= len(cleaned):
            break
        start += step
    return chunks


def cosine_similarity(vec_a: List[float], vec_b: List[float]) -> float:
    if len(vec_a) != len(vec_b):
        return 0.0

    dot = sum(a * b for a, b in zip(vec_a, vec_b))
    norm_a = math.sqrt(sum(a * a for a in vec_a))
    norm_b = math.sqrt(sum(b * b for b in vec_b))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


def looks_like_text(decoded_text: str) -> bool:
    if not decoded_text:
        return False
    sample = decoded_text[:2000]
    printable = sum(ch.isprintable() or ch in "\n\r\t" for ch in sample)
    ratio = printable / max(len(sample), 1)
    return ratio >= 0.85


def normalize_folder_color(color: Optional[str]) -> str:
    if color is None:
        return DEFAULT_FOLDER_COLOR

    normalized = color.strip().lower()
    if normalized not in FOLDER_COLOR_OPTIONS:
        raise ValueError("Invalid folder color")
    return normalized


async def embed_texts(inputs: List[str], model_name: str = EMBEDDING_MODEL) -> List[List[float]]:
    if not inputs:
        return []

    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(
            f"{OLLAMA_BASE_URL}/api/embed",
            json={"model": model_name, "input": inputs},
        )

    if response.status_code != 200:
        raise RuntimeError(f"Embedding request failed with status {response.status_code}")

    data = response.json()
    embeddings = data.get("embeddings")
    if not isinstance(embeddings, list):
        raise RuntimeError("Embedding response missing 'embeddings'")
    return embeddings


async def fetch_ollama_model_names() -> List[str]:
    async with httpx.AsyncClient(timeout=8.0) as client:
        response = await client.get(f"{OLLAMA_BASE_URL}/api/tags")

    if response.status_code != 200:
        raise RuntimeError(f"Ollama tags request failed with status {response.status_code}")

    payload = response.json()
    models = payload.get("models")
    if not isinstance(models, list):
        raise RuntimeError("Invalid Ollama tags response")

    names: List[str] = []
    for model in models:
        if isinstance(model, dict):
            name = str(model.get("name", "")).strip()
            if name:
                names.append(name)

    return sorted(set(names))


def is_ollama_installed() -> bool:
    if shutil.which("ollama"):
        return True

    if sys.platform == "darwin":
        app_candidates = [
            Path("/Applications/Ollama.app"),
            Path.home() / "Applications" / "Ollama.app",
        ]
        return any(path.exists() for path in app_candidates)

    return False


async def is_ollama_running() -> bool:
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            response = await client.get(f"{OLLAMA_BASE_URL}/api/version")
        return response.status_code == 200
    except (httpx.ConnectError, httpx.TimeoutException, httpx.HTTPError):
        return False


async def install_ollama_macos() -> Tuple[bool, str]:
    if sys.platform != "darwin":
        return False, "Automatic Ollama install is only supported on macOS."

    brew_bin = shutil.which("brew")
    if not brew_bin:
        return (
            False,
            "Homebrew is required for automatic Ollama install. Install Homebrew and try again.",
        )

    def _run_install() -> Tuple[bool, str]:
        try:
            result = subprocess.run(
                [brew_bin, "install", "--cask", "ollama"],
                check=False,
                capture_output=True,
                text=True,
            )
        except Exception as error:
            return False, f"Failed to run Homebrew install: {type(error).__name__}: {error}"

        if result.returncode != 0:
            output = (result.stderr or result.stdout or "").strip()
            if output:
                output = output.splitlines()[-1].strip()
            else:
                output = f"exit code {result.returncode}"
            return False, f"Homebrew could not install Ollama: {output}"

        return True, ""

    return await asyncio.to_thread(_run_install)


def try_start_ollama() -> bool:
    try:
        if sys.platform == "darwin":
            open_result = subprocess.run(
                ["open", "-a", "Ollama"],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            if open_result.returncode == 0:
                return True
    except Exception as error:
        print(f"Ollama app start warning: {type(error).__name__}: {error}", file=sys.stderr)

    ollama_bin = shutil.which("ollama")
    if not ollama_bin:
        return False

    try:
        subprocess.Popen(
            [ollama_bin, "serve"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        return True
    except Exception as error:
        print(f"Ollama serve start warning: {type(error).__name__}: {error}", file=sys.stderr)
        return False


async def wait_for_ollama_ready(max_wait_seconds: int = 25) -> bool:
    for _ in range(max_wait_seconds):
        if await is_ollama_running():
            return True
        await asyncio.sleep(1)
    return False


async def pull_ollama_model(model_name: str) -> None:
    async with httpx.AsyncClient(timeout=1800.0) as client:
        response = await client.post(
            f"{OLLAMA_BASE_URL}/api/pull",
            json={"name": model_name, "stream": False},
        )

    if response.status_code != 200:
        detail = ""
        try:
            payload = response.json()
            if isinstance(payload, dict):
                detail = str(payload.get("error", "")).strip()
        except ValueError:
            detail = response.text.strip()

        if not detail:
            detail = f"status {response.status_code}"

        raise RuntimeError(f"Failed to pull model '{model_name}': {detail}")


async def stream_pull_ollama_model_progress(
    model_name: str,
) -> AsyncGenerator[Dict[str, Any], None]:
    async with httpx.AsyncClient(timeout=1800.0) as client:
        async with client.stream(
            "POST",
            f"{OLLAMA_BASE_URL}/api/pull",
            json={"name": model_name, "stream": True},
        ) as response:
            if response.status_code != 200:
                detail = ""
                try:
                    payload = response.json()
                    if isinstance(payload, dict):
                        detail = str(payload.get("error", "")).strip()
                except Exception:
                    detail = ""
                if not detail:
                    detail = f"status {response.status_code}"
                raise RuntimeError(f"Failed to pull model '{model_name}': {detail}")

            async for line in response.aiter_lines():
                if not line.strip():
                    continue
                try:
                    data = json.loads(line)
                except json.JSONDecodeError:
                    continue

                if not isinstance(data, dict):
                    continue

                if data.get("error"):
                    raise RuntimeError(
                        f"Failed to pull model '{model_name}': {data['error']}"
                    )

                yield {
                    "status": str(data.get("status", "")).strip(),
                    "completed": data.get("completed"),
                    "total": data.get("total"),
                }


async def build_setup_prerequisites_status() -> Dict[str, Any]:
    installed = is_ollama_installed()
    running = await is_ollama_running() if installed else False

    available_models: List[str] = []
    if running:
        try:
            available_models = await fetch_ollama_model_names()
        except Exception as error:
            print(
                f"Model list warning during setup status: {type(error).__name__}: {error}",
                file=sys.stderr,
            )

    available_set = {
        normalize_ollama_model_name(model_name)
        for model_name in available_models
        if model_name.strip()
    }
    missing_models = [
        model_name
        for model_name in REQUIRED_VESTA_MODELS
        if normalize_ollama_model_name(model_name) not in available_set
    ]

    return {
        "ollama_installed": installed,
        "ollama_running": running,
        "required_models": REQUIRED_VESTA_MODELS,
        "available_models": available_models,
        "missing_models": missing_models,
        "ready": installed and running and not missing_models,
    }


def resolve_target_models(
    status: Dict[str, Any], requested_models: Optional[List[str]]
) -> List[str]:
    available_set = {
        normalize_ollama_model_name(str(model_name))
        for model_name in status.get("available_models", [])
        if str(model_name).strip()
    }
    if not requested_models:
        return list(status.get("missing_models", []))

    normalized: List[str] = []
    seen: set[str] = set()
    for raw_model in requested_models:
        model_name = raw_model.strip()
        if not model_name or model_name in seen:
            continue
        seen.add(model_name)
        normalized.append(model_name)

    return [
        model_name
        for model_name in normalized
        if normalize_ollama_model_name(model_name) not in available_set
    ]


async def embed_in_batches(chunks: List[str], batch_size: int = 24) -> List[List[float]]:
    vectors: List[List[float]] = []
    for start in range(0, len(chunks), batch_size):
        batch = chunks[start : start + batch_size]
        vectors.extend(await embed_texts(batch))
    return vectors


def extract_pdf_text(content: bytes) -> str:
    """Extract text from PDF."""
    try:
        import PyPDF2

        pdf_file = io.BytesIO(content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text() or ""
            text += page_text + "\n"
        return text.strip()
    except ImportError:
        return "[PDF processing unavailable - PyPDF2 not installed]"
    except Exception as e:
        return f"[Error extracting PDF: {str(e)}]"


def extract_docx_text(content: bytes) -> str:
    """Extract text from DOCX."""
    try:
        import docx

        doc_file = io.BytesIO(content)
        doc = docx.Document(doc_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text.strip()
    except ImportError:
        return "[DOCX processing unavailable - python-docx not installed]"
    except Exception as e:
        return f"[Error extracting DOCX: {str(e)}]"


def extract_csv_text(content: bytes) -> str:
    """Extract text from CSV as markdown table."""
    try:
        csv_file = io.StringIO(content.decode("utf-8", errors="ignore"))
        reader = csv.reader(csv_file)
        rows = list(reader)

        if not rows:
            return "[Empty CSV file]"

        text = "| " + " | ".join(rows[0]) + " |\n"
        text += "| " + " | ".join(["---"] * len(rows[0])) + " |\n"
        for row in rows[1:]:
            text += "| " + " | ".join(row) + " |\n"
        return text
    except Exception as e:
        return f"[Error extracting CSV: {str(e)}]"


def extract_excel_text(content: bytes) -> str:
    """Extract text from Excel files."""
    try:
        import openpyxl

        excel_file = io.BytesIO(content)
        workbook = openpyxl.load_workbook(excel_file, read_only=True)

        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n\n"

            rows = []
            for row in sheet.iter_rows(values_only=True, max_row=100):
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if rows:
                text += "| " + " | ".join(rows[0]) + " |\n"
                text += "| " + " | ".join(["---"] * len(rows[0])) + " |\n"
                for row in rows[1:]:
                    text += "| " + " | ".join(row) + " |\n"

        return text.strip()
    except ImportError:
        return "[Excel processing unavailable - openpyxl not installed]"
    except Exception as e:
        return f"[Error extracting Excel: {str(e)}]"


def extract_text_for_knowledge(filename: str, content: bytes) -> Tuple[Optional[str], Optional[str]]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if extension == "pdf":
        extracted = extract_pdf_text(content)
        return (extracted, None) if extracted else (None, "No text extracted from PDF")
    if extension in ["docx", "doc"]:
        extracted = extract_docx_text(content)
        return (extracted, None) if extracted else (None, "No text extracted from document")
    if extension == "csv":
        extracted = extract_csv_text(content)
        return (extracted, None) if extracted else (None, "No text extracted from CSV")
    if extension == "txt":
        extracted = content.decode("utf-8", errors="ignore").strip()
        return (extracted, None) if extracted else (None, "Text file is empty")
    if extension in ["xlsx", "xls"]:
        extracted = extract_excel_text(content)
        return (extracted, None) if extracted else (None, "No text extracted from spreadsheet")

    decoded = content.decode("utf-8", errors="ignore").strip()
    if decoded and looks_like_text(decoded):
        return decoded, None

    return None, "Unsupported or non-text binary file"


async def retrieve_knowledge_context(
    query: str, folder_id: Optional[str] = None
) -> Tuple[str, List[Dict[str, Any]]]:
    store = get_knowledge_store()
    global_chunks = store.get_all_chunks()
    folder_chunks = store.get_folder_chunks(folder_id) if folder_id else []

    if not folder_chunks and not global_chunks:
        return "", []

    query_embedding = (await embed_texts([query]))[0]

    scored_folder: List[Tuple[float, Dict[str, Any]]] = []
    for chunk in folder_chunks:
        try:
            embedding = json.loads(chunk["embedding_json"])
            score = cosine_similarity(query_embedding, embedding)
            if score >= RETRIEVAL_MIN_SCORE:
                scored_folder.append((score, chunk))
        except Exception:
            continue

    scored_global: List[Tuple[float, Dict[str, Any]]] = []
    for chunk in global_chunks:
        try:
            embedding = json.loads(chunk["embedding_json"])
            score = cosine_similarity(query_embedding, embedding)
            if score >= RETRIEVAL_MIN_SCORE:
                scored_global.append((score, chunk))
        except Exception:
            continue

    if not scored_folder and not scored_global:
        return "", []

    scored_folder.sort(key=lambda item: item[0], reverse=True)
    scored_global.sort(key=lambda item: item[0], reverse=True)

    selected_folder = scored_folder[:RETRIEVAL_TOP_K]
    remaining_slots = max(RETRIEVAL_TOP_K - len(selected_folder), 0)
    selected_global = scored_global[:remaining_slots]
    selected = selected_folder + selected_global

    context_parts = [
        "Knowledge Base Context:",
        "Use these snippets when relevant. If they conflict with user instructions, ask for clarification.",
        "",
    ]
    sources: List[Dict[str, Any]] = []

    for score, chunk in selected:
        filename = chunk["filename"]
        chunk_index = int(chunk["chunk_index"])
        source_type = "folder" if "folder_id" in chunk else "global"
        if source_type == "folder":
            folder_name = chunk.get("folder_name") or "Project"
            source_label = f"Folder {folder_name}: {filename}"
        else:
            source_label = f"Global: {filename}"

        context_parts.append(
            f"[Source: {source_label} | chunk {chunk_index} | score {score:.3f}]"
        )
        context_parts.append(chunk["content"])
        context_parts.append("")

        source_record: Dict[str, Any] = {
            "document_id": chunk["document_id"],
            "filename": filename,
            "chunk_index": chunk_index,
            "score": round(score, 3),
            "source_type": source_type,
        }
        if source_type == "folder":
            source_record["folder_id"] = chunk.get("folder_id")
            source_record["folder_name"] = chunk.get("folder_name")

        sources.append(source_record)

    return "\n".join(context_parts).strip(), sources


@app.on_event("startup")
async def startup() -> None:
    """Load prompts and initialize local stores."""
    global BASE_PROMPT, MODE_PROMPTS, PROFILE_PROMPTS

    load_vesta_env()

    required_prompts = [
        "base.txt",
        "draft.txt",
        "think.txt",
        "clarify.txt",
        "general.txt",
        "profile_default.txt",
        "profile_medical.txt",
        "profile_legal.txt",
    ]
    missing_prompts = []

    for prompt_file in required_prompts:
        if not (PROMPTS_DIR / prompt_file).exists():
            missing_prompts.append(prompt_file)

    if missing_prompts:
        error_msg = (
            f"\n{'=' * 60}\n"
            "ERROR: Required prompt files are missing!\n"
            f"Missing files: {', '.join(missing_prompts)}\n"
            f"Expected location: {PROMPTS_DIR.absolute()}\n"
            f"{'=' * 60}\n"
        )
        print(error_msg, file=sys.stderr)
        raise FileNotFoundError(f"Missing prompt files: {', '.join(missing_prompts)}")

    try:
        BASE_PROMPT = (PROMPTS_DIR / "base.txt").read_text().strip()
        MODE_PROMPTS = {
            "draft": (PROMPTS_DIR / "draft.txt").read_text().strip(),
            "think": (PROMPTS_DIR / "think.txt").read_text().strip(),
            "clarify": (PROMPTS_DIR / "clarify.txt").read_text().strip(),
            "general": (PROMPTS_DIR / "general.txt").read_text().strip(),
        }
        PROFILE_PROMPTS = {
            "default": (PROMPTS_DIR / "profile_default.txt").read_text().strip(),
            "medical": (PROMPTS_DIR / "profile_medical.txt").read_text().strip(),
            "legal": (PROMPTS_DIR / "profile_legal.txt").read_text().strip(),
        }

        get_knowledge_store()
        print(
            f"Successfully loaded {len(MODE_PROMPTS) + len(PROFILE_PROMPTS) + 1} prompt files from {PROMPTS_DIR}"
        )
    except Exception as e:
        print(f"Error during startup initialization: {e}", file=sys.stderr)
        raise


async def route_to_model(
    message: str,
    mode: str,
    history: List[ChatMessage],
    last_model_used: Optional[str] = None,
) -> RoutingDecision:
    """
    VESTA-compliant routing with coherence framework analysis.

    Returns: RoutingDecision with full audit trail
    """
    start_time = time.time()

    history_dicts = [{"role": msg.role, "content": msg.content} for msg in history]

    signals = analyze_message_signals(message, mode, history_dicts)
    task_context = analyze_task_context(history_dicts, message)

    upgraded_model = should_upgrade_model(message, history_dicts, last_model_used)
    if upgraded_model:
        _latency_ms = (time.time() - start_time) * 1000
        return RoutingDecision(
            model=upgraded_model,
            method="refinement_upgrade",
            reasoning=f"User requested refinement, upgrading from {last_model_used}",
            signals=signals,
            task_context=task_context,
            confidence=0.95,
            fallback_used=False,
        )

    decision = fast_route(signals, mode, task_context)
    if decision:
        _latency_ms = (time.time() - start_time) * 1000
        return decision

    try:
        decision = await llm_route(message, mode, signals, task_context, history)
        _latency_ms = (time.time() - start_time) * 1000
        return decision
    except Exception as e:
        _latency_ms = (time.time() - start_time) * 1000
        fallback_model = get_fallback_model(mode)

        log_routing_error(message, mode, e, fallback_model)

        return RoutingDecision(
            model=fallback_model,
            method="fallback",
            reasoning="Routing error, using mode-based fallback",
            signals=signals,
            task_context=task_context,
            confidence=0.5,
            fallback_used=True,
        )


async def llm_route(
    message: str,
    mode: str,
    signals: Any,
    task_context: Any,
    history: List[ChatMessage],
) -> RoutingDecision:
    """
    LLM-based routing for ambiguous cases.
    Enhanced with coherence framework signals.
    """
    configured_models = get_knowledge_store().get_model_names()
    lite_model_name = configured_models["lite"]
    general_model_name = configured_models["general"]
    deep_model_name = configured_models["deep"]

    routing_prompt = f"""Analyze this user query and determine which AI model should handle it.

Available model profiles:
- lite ({lite_model_name}): For simple, straightforward questions, quick clarifications, basic information retrieval
- general ({general_model_name}): For standard tasks, moderate complexity, general conversation, typical problem-solving
- deep ({deep_model_name}): For complex reasoning, deep analysis, nuanced thinking, difficult problems requiring extensive reasoning

User's selected mode: {mode}
Message context: {len(history)} previous messages

Coherence Analysis:
- Energy (computational complexity): {signals.energy:.2f}
- Information (context integration): {signals.information:.2f}
- Connection (relational depth): {signals.connection:.2f}
- Noise tolerance: {signals.noise_tolerance:.2f}

Task Context:
- Is continuation: {task_context.is_continuation}
- Task depth: {task_context.depth}
- Requires consistency: {task_context.requires_consistency}
- Complexity trend: {task_context.complexity_trend}

User query: {message}

Consider all factors and respond with ONLY a JSON object:
{{"model": "lite|general|deep", "reasoning": "brief explanation", "confidence": 0.0-1.0}}"""

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json={
                "model": general_model_name,
                "prompt": routing_prompt,
                "temperature": 0.3,
                "stream": False,
            },
        )

    if response.status_code == 200:
        result = response.json()
        response_text = result.get("response", "").strip()

        start = response_text.find("{")
        end = response_text.rfind("}") + 1
        if start >= 0 and end > start:
            json_str = response_text[start:end]
            decision_data = json.loads(json_str)

            selected_model = decision_data.get("model", "general")
            reasoning = decision_data.get("reasoning", "LLM routing decision")
            confidence = decision_data.get("confidence", 0.7)

            if selected_model in ["lite", "general", "deep"]:
                return RoutingDecision(
                    model=selected_model,
                    method="llm",
                    reasoning=reasoning,
                    signals=signals,
                    task_context=task_context,
                    confidence=float(confidence),
                    fallback_used=False,
                )

    raise ValueError("Invalid LLM routing response")


def get_fallback_model(mode: str) -> str:
    """Get fallback model based on mode."""
    fallbacks = {
        "think": "deep",
        "draft": "general",
        "clarify": "general",
        "general": "general",
    }
    return fallbacks.get(mode, "general")


def build_conversation_context(
    messages: List[ChatMessage],
    profile_prompt: str,
    mode_prompt: str,
    current_message: str,
    knowledge_context: str = "",
) -> str:
    """Build the full conversation context including history and optional retrieved knowledge."""
    context_parts = [BASE_PROMPT, "", profile_prompt, "", mode_prompt, ""]

    if knowledge_context:
        context_parts.extend([knowledge_context, ""])

    if messages:
        context_parts.append("Previous conversation:")
        for msg in messages[-10:]:
            prefix = "User: " if msg.role == "user" else "Assistant: "
            context_parts.append(f"{prefix}{msg.content}")
        context_parts.append("")

    context_parts.append(f"User: {current_message}")

    return "\n".join(context_parts)


async def stream_ollama_response(
    model_name: str,
    prompt: str,
    sources: Optional[List[Dict[str, Any]]] = None,
):
    """Stream response from Ollama and emit source metadata first."""
    metadata_payload = {"metadata": {"sources": sources or []}}
    yield f"data: {json.dumps(metadata_payload)}\n\n"

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            async with client.stream(
                "POST",
                f"{OLLAMA_BASE_URL}/api/generate",
                json={"model": model_name, "prompt": prompt, "stream": True},
            ) as response:
                if response.status_code != 200:
                    yield f"data: {json.dumps({'error': 'Ollama service unavailable'})}\n\n"
                    return

                async for line in response.aiter_lines():
                    if line.strip():
                        try:
                            data = json.loads(line)
                            if "response" in data:
                                chunk = {
                                    "content": data["response"],
                                    "done": data.get("done", False),
                                }
                                yield f"data: {json.dumps(chunk)}\n\n"

                                if data.get("done"):
                                    break
                        except json.JSONDecodeError:
                            continue

    except httpx.ConnectError:
        yield f"data: {json.dumps({'error': 'Cannot connect to Ollama'})}\n\n"
    except Exception as e:
        print(f"Streaming error: {e}", file=sys.stderr)
        yield f"data: {json.dumps({'error': 'Streaming failed'})}\n\n"


@app.post("/chat")
async def chat(request: ChatRequest):
    """
    Handle chat requests with context and streaming.
    Supports manual model selection or auto-routing with coherence framework.
    """
    try:
        start_time = time.time()
        routing_decision = None
        consistency_enforced = False
        selected_profile = normalize_assistant_profile(request.profile)

        if request.model == "auto":
            routing_decision = await route_to_model(
                request.message,
                request.mode,
                request.messages,
                request.last_model_used,
            )
            selected_model_key = routing_decision.model

            original_model = selected_model_key
            selected_model_key, was_upgraded = enforce_model_consistency(
                selected_model_key,
                [{"role": msg.role, "content": msg.content} for msg in request.messages],
                request.last_model_used,
            )

            if was_upgraded:
                consistency_enforced = True
                log_model_consistency_event(
                    original_model,
                    selected_model_key,
                    "Prevented mid-task downgrade",
                    len([m for m in request.messages if m.role == "user"]),
                )

            if selected_profile in {"medical", "legal"} and selected_model_key == "lite":
                selected_model_key = "general"
        else:
            selected_model_key = request.model if request.model != "auto" else "general"

        configured_model_names = get_knowledge_store().get_model_names()
        model_name = configured_model_names.get(
            selected_model_key,
            configured_model_names["general"],
        )

        if routing_decision:
            latency_ms = (time.time() - start_time) * 1000
            log_routing_decision(
                message=request.message,
                mode=request.mode,
                history_depth=len([m for m in request.messages if m.role == "user"]),
                signals={
                    "energy": routing_decision.signals.energy,
                    "information": routing_decision.signals.information,
                    "connection": routing_decision.signals.connection,
                    "noise_tolerance": routing_decision.signals.noise_tolerance,
                },
                task_context={
                    "is_continuation": routing_decision.task_context.is_continuation,
                    "is_new_task": routing_decision.task_context.is_new_task,
                    "depth": routing_decision.task_context.depth,
                    "requires_consistency": routing_decision.task_context.requires_consistency,
                    "complexity_trend": routing_decision.task_context.complexity_trend,
                    "task_type": routing_decision.task_context.task_type,
                },
                routing_method=routing_decision.method,
                selected_model=selected_model_key,
                reasoning=routing_decision.reasoning,
                confidence=routing_decision.confidence,
                fallback_used=routing_decision.fallback_used,
                latency_ms=latency_ms,
                last_model_used=request.last_model_used,
                consistency_enforced=consistency_enforced,
            )

        effective_folder_id = request.folder_id
        if effective_folder_id is None and request.conversation_id:
            effective_folder_id = get_knowledge_store().get_conversation_folder_id(
                request.conversation_id
            )

        knowledge_context = ""
        knowledge_sources: List[Dict[str, Any]] = []
        try:
            knowledge_context, knowledge_sources = await retrieve_knowledge_context(
                request.message, effective_folder_id
            )
        except Exception as retrieval_error:
            print(
                f"Knowledge retrieval warning: {type(retrieval_error).__name__}: {retrieval_error}",
                file=sys.stderr,
            )

        weather_context = ""
        weather_sources: List[Dict[str, Any]] = []
        if is_weather_intent(request.message):
            try:
                weather_context, weather_sources = await retrieve_weather_context(request.message)
            except Exception as weather_error:
                print(
                    f"Weather context warning: {type(weather_error).__name__}: {weather_error}",
                    file=sys.stderr,
                )

        context_blocks = [block for block in [knowledge_context, weather_context] if block]
        combined_context = "\n\n".join(context_blocks)
        combined_sources = [*knowledge_sources, *weather_sources]

        profile_prompt = PROFILE_PROMPTS.get(selected_profile, PROFILE_PROMPTS.get("default", ""))
        mode_prompt = MODE_PROMPTS.get(request.mode, "")
        full_prompt = build_conversation_context(
            request.messages,
            profile_prompt,
            mode_prompt,
            request.message,
            combined_context,
        )

        response = StreamingResponse(
            stream_ollama_response(model_name, full_prompt, combined_sources),
            media_type="text/event-stream",
        )

        response.headers["X-Selected-Model"] = selected_model_key
        if routing_decision:
            response.headers["X-Routing-Method"] = routing_decision.method
            response.headers["X-Routing-Confidence"] = str(round(routing_decision.confidence, 2))

        return response

    except Exception as e:
        print(f"Unexpected error in chat endpoint: {type(e).__name__}: {e}", file=sys.stderr)
        raise HTTPException(
            status_code=500,
            detail="An error occurred processing your request.",
        )


@app.post("/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    """
    Process uploaded files and extract text content.
    VESTA-compliant: Files are processed immediately and not stored.
    Returns extracted text for session-scoped context only.
    """
    extracted_content = []

    for file in files:
        try:
            content = await file.read()
            file_type = file.filename.split(".")[-1].lower() if "." in file.filename else ""

            if len(content) > MAX_CHAT_UPLOAD_SIZE:
                extracted_content.append(
                    {"filename": file.filename, "error": "File too large (max 10MB)"}
                )
                continue

            if file_type == "pdf":
                text = extract_pdf_text(content)
            elif file_type in ["docx", "doc"]:
                text = extract_docx_text(content)
            elif file_type in ["csv"]:
                text = extract_csv_text(content)
            elif file_type in ["txt"]:
                text = content.decode("utf-8", errors="ignore")
            elif file_type in ["xlsx", "xls"]:
                text = extract_excel_text(content)
            else:
                text = f"[Unsupported file type: {file_type}]"

            extracted_content.append(
                {
                    "filename": file.filename,
                    "content": text[:50000],
                    "size": len(content),
                }
            )
        except Exception as e:
            print(f"Error processing file {file.filename}: {e}", file=sys.stderr)
            extracted_content.append({"filename": file.filename, "error": str(e)})

    return {"files": extracted_content}


@app.post("/knowledge/files")
async def upload_knowledge_files(files: List[UploadFile] = File(...)):
    """Persist files for retrieval-augmented chat."""
    store = get_knowledge_store()
    results: List[Dict[str, Any]] = []

    for file in files:
        try:
            content = await file.read()
            if len(content) > MAX_KNOWLEDGE_UPLOAD_SIZE:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "error",
                        "reason": "File too large (max 25MB)",
                    }
                )
                continue

            text, extraction_error = extract_text_for_knowledge(file.filename, content)
            if extraction_error:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": extraction_error,
                    }
                )
                continue

            if not text:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": "No text extracted from file",
                    }
                )
                continue

            trimmed_text = text[:MAX_KNOWLEDGE_TEXT_CHARS]
            chunks = chunk_text(trimmed_text)
            if not chunks:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": "No meaningful text chunks could be produced",
                    }
                )
                continue

            content_hash = hashlib.sha256(content).hexdigest()
            duplicate = store.get_document_by_hash(content_hash)
            if duplicate:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "duplicate",
                        "reason": "Document with identical content already indexed",
                        "document": duplicate,
                    }
                )
                continue

            embeddings = await embed_in_batches(chunks)
            document = store.insert_document_with_chunks(
                filename=file.filename,
                content_hash=content_hash,
                size_bytes=len(content),
                mime_type=file.content_type,
                chunks=chunks,
                embeddings=embeddings,
            )

            results.append(
                {
                    "filename": file.filename,
                    "status": "indexed",
                    "document": document,
                }
            )
        except Exception as e:
            print(f"Knowledge ingest error for {file.filename}: {e}", file=sys.stderr)
            results.append(
                {
                    "filename": file.filename,
                    "status": "error",
                    "reason": f"{type(e).__name__}: {e}",
                }
            )

    return {"results": results}


@app.get("/knowledge/files")
async def list_knowledge_files():
    store = get_knowledge_store()
    return {"documents": store.list_documents()}


@app.delete("/knowledge/files/{document_id}")
async def delete_knowledge_file(document_id: str):
    store = get_knowledge_store()
    result = store.delete_document(document_id)
    if not result["deleted"]:
        raise HTTPException(status_code=404, detail="Document not found")

    return {
        "deleted": True,
        "document_id": document_id,
        "chunk_count": result["chunk_count"],
    }


@app.get("/folders")
async def list_folders():
    store = get_knowledge_store()
    return {"folders": store.list_folders()}


@app.post("/folders")
async def create_folder(request: FolderCreateRequest):
    store = get_knowledge_store()
    name = request.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Folder name is required")
    try:
        color = normalize_folder_color(request.color)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid folder color")

    try:
        folder = store.create_folder(name, color)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Folder name already exists")

    return {"folder": folder}


@app.patch("/folders/{folder_id}")
async def rename_folder(folder_id: str, request: FolderUpdateRequest):
    store = get_knowledge_store()
    name = request.name.strip() if request.name is not None else None
    if request.name is not None and not name:
        raise HTTPException(status_code=400, detail="Folder name is required")
    if request.name is None and request.color is None:
        raise HTTPException(status_code=400, detail="No folder updates provided")

    color: Optional[str] = None
    if request.color is not None:
        try:
            color = normalize_folder_color(request.color)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid folder color")

    try:
        folder = store.update_folder(folder_id, name=name, color=color)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Folder name already exists")

    if folder is None:
        raise HTTPException(status_code=404, detail="Folder not found")

    return {"folder": folder}


@app.delete("/folders/{folder_id}")
async def delete_folder(folder_id: str):
    store = get_knowledge_store()
    result = store.delete_folder(folder_id)
    if not result["deleted"]:
        raise HTTPException(status_code=404, detail="Folder not found")

    return {
        "deleted": True,
        "folder_id": folder_id,
        "conversations_deleted": result["conversations_deleted"],
        "documents_deleted": result["documents_deleted"],
        "chunks_deleted": result["chunks_deleted"],
    }


@app.get("/folders/{folder_id}/files")
async def list_folder_knowledge_files(folder_id: str):
    store = get_knowledge_store()
    if store.get_folder(folder_id) is None:
        raise HTTPException(status_code=404, detail="Folder not found")
    return {"documents": store.list_folder_documents(folder_id)}


@app.post("/folders/{folder_id}/files")
async def upload_folder_knowledge_files(folder_id: str, files: List[UploadFile] = File(...)):
    store = get_knowledge_store()
    if store.get_folder(folder_id) is None:
        raise HTTPException(status_code=404, detail="Folder not found")

    results: List[Dict[str, Any]] = []

    for file in files:
        try:
            content = await file.read()
            if len(content) > MAX_KNOWLEDGE_UPLOAD_SIZE:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "error",
                        "reason": "File too large (max 25MB)",
                    }
                )
                continue

            text, extraction_error = extract_text_for_knowledge(file.filename, content)
            if extraction_error:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": extraction_error,
                    }
                )
                continue

            if not text:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": "No text extracted from file",
                    }
                )
                continue

            trimmed_text = text[:MAX_KNOWLEDGE_TEXT_CHARS]
            chunks = chunk_text(trimmed_text)
            if not chunks:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "unsupported",
                        "reason": "No meaningful text chunks could be produced",
                    }
                )
                continue

            content_hash = hashlib.sha256(content).hexdigest()
            duplicate = store.get_folder_document_by_hash(folder_id, content_hash)
            if duplicate:
                results.append(
                    {
                        "filename": file.filename,
                        "status": "duplicate",
                        "reason": "Document with identical content already indexed in this folder",
                        "document": duplicate,
                    }
                )
                continue

            embeddings = await embed_in_batches(chunks)
            document = store.insert_folder_document_with_chunks(
                folder_id=folder_id,
                filename=file.filename,
                content_hash=content_hash,
                size_bytes=len(content),
                mime_type=file.content_type,
                chunks=chunks,
                embeddings=embeddings,
            )

            results.append(
                {
                    "filename": file.filename,
                    "status": "indexed",
                    "document": document,
                }
            )
        except Exception as e:
            print(
                f"Folder knowledge ingest error for {file.filename}: {e}",
                file=sys.stderr,
            )
            results.append(
                {
                    "filename": file.filename,
                    "status": "error",
                    "reason": f"{type(e).__name__}: {e}",
                }
            )

    return {"results": results}


@app.delete("/folders/{folder_id}/files/{document_id}")
async def delete_folder_knowledge_file(folder_id: str, document_id: str):
    store = get_knowledge_store()
    if store.get_folder(folder_id) is None:
        raise HTTPException(status_code=404, detail="Folder not found")

    result = store.delete_folder_document(folder_id, document_id)
    if not result["deleted"]:
        raise HTTPException(status_code=404, detail="Document not found")

    return {
        "deleted": True,
        "document_id": document_id,
        "chunk_count": result["chunk_count"],
    }


@app.get("/conversations")
async def list_conversations():
    store = get_knowledge_store()
    return {"conversations": store.list_conversations()}


@app.post("/conversations")
async def create_conversation(request: ConversationCreateRequest):
    store = get_knowledge_store()
    try:
        conversation = store.create_conversation(request.folder_id)
    except ValueError as error:
        raise HTTPException(status_code=404, detail=str(error))

    return {"conversation": conversation}


@app.get("/conversations/{conversation_id}")
async def get_conversation(conversation_id: str):
    store = get_knowledge_store()
    conversation = store.get_conversation(conversation_id)
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    return {
        "conversation": conversation,
        "messages": store.get_conversation_messages(conversation_id),
    }


@app.patch("/conversations/{conversation_id}")
async def update_conversation(conversation_id: str, request: ConversationUpdateRequest):
    store = get_knowledge_store()
    payload = request.model_dump(exclude_unset=True)

    title = payload.get("title")
    if title is not None:
        title = title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="Conversation title is required")

    set_folder = "folder_id" in payload
    folder_id = payload.get("folder_id")

    try:
        conversation = store.update_conversation(
            conversation_id,
            title=title,
            set_folder=set_folder,
            folder_id=folder_id,
        )
    except ValueError as error:
        raise HTTPException(status_code=404, detail=str(error))

    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    return {"conversation": conversation}


@app.delete("/conversations/{conversation_id}")
async def delete_conversation(conversation_id: str):
    store = get_knowledge_store()
    result = store.delete_conversation(conversation_id)
    if not result["deleted"]:
        raise HTTPException(status_code=404, detail="Conversation not found")

    return {"deleted": True, "conversation_id": conversation_id}


@app.post("/conversations/{conversation_id}/turns")
async def append_conversation_turn(conversation_id: str, request: ConversationTurnRequest):
    user_message = request.user_message.strip()
    assistant_message = request.assistant_message.strip()
    if not user_message:
        raise HTTPException(status_code=400, detail="User message is required")
    if not assistant_message:
        raise HTTPException(status_code=400, detail="Assistant message is required")

    store = get_knowledge_store()
    conversation = store.append_turn(
        conversation_id,
        user_message=user_message,
        assistant_message=assistant_message,
        model_used=request.model_used,
        sources=request.sources,
    )
    if conversation is None:
        raise HTTPException(status_code=404, detail="Conversation not found")

    return {"saved": True, "conversation": conversation}


@app.get("/setup/prerequisites")
async def get_setup_prerequisites():
    return await build_setup_prerequisites_status()


@app.post("/setup/prerequisites")
async def run_setup_prerequisites(request: SetupPrerequisitesRequest):
    if not request.approved:
        raise HTTPException(
            status_code=400,
            detail="User approval is required before running setup.",
        )

    store = get_knowledge_store()
    run_id = store.create_setup_run(request.models)

    def log_event(
        event_type: str,
        *,
        message: Optional[str] = None,
        model_name: Optional[str] = None,
        payload: Optional[Dict[str, Any]] = None,
    ) -> None:
        store.append_setup_run_event(
            run_id,
            event_type,
            message=message,
            model_name=model_name,
            payload=payload,
        )

    installed_ollama = False
    started_ollama = False
    pulled_models: List[str] = []
    failed_models: List[Dict[str, str]] = []

    try:
        log_event("setup_start", payload={"requested_models": request.models or []})

        status = await build_setup_prerequisites_status()
        log_event("status", payload=status)

        if not status["ollama_installed"]:
            log_event("install_start")
            install_success, install_message = await install_ollama_macos()
            if not install_success:
                raise RuntimeError(install_message)
            installed_ollama = True
            log_event("install_done")
            status = await build_setup_prerequisites_status()
            log_event("status", payload=status)
            if not status["ollama_installed"]:
                raise RuntimeError(
                    "Ollama installation completed but Ollama is still not detected."
                )

        if not status["ollama_running"]:
            log_event("start_ollama")
            started_ollama = try_start_ollama()
            if not started_ollama:
                raise RuntimeError(
                    "Could not start Ollama automatically. Start Ollama and try again."
                )
            if not await wait_for_ollama_ready():
                raise RuntimeError("Ollama was started but did not become ready in time.")
            log_event("start_ollama_done")
            status = await build_setup_prerequisites_status()
            log_event("status", payload=status)

        target_models = resolve_target_models(status, request.models)
        log_event("target_models", payload={"target_models": target_models})

        for model_name in target_models:
            log_event("pull_start", model_name=model_name)
            try:
                await pull_ollama_model(model_name)
                pulled_models.append(model_name)
                log_event("pull_done", model_name=model_name)
            except Exception as error:
                failure_message = str(error)
                failed_models.append({"model": model_name, "error": failure_message})
                log_event("pull_error", model_name=model_name, message=failure_message)

        final_status = await build_setup_prerequisites_status()
        ready = final_status["ready"] and not failed_models
        log_event("complete", payload={"ready": ready, "status": final_status})
        store.finish_setup_run(
            run_id,
            success=ready,
            installed_ollama=installed_ollama,
            started_ollama=started_ollama,
            pulled_models=pulled_models,
            failed_models=failed_models,
        )
        return {
            "run_id": run_id,
            "approved": True,
            "installed_ollama": installed_ollama,
            "started_ollama": started_ollama,
            "pulled_models": pulled_models,
            "failed_models": failed_models,
            "status": final_status,
            "ready": ready,
        }
    except Exception as error:
        log_event("error", message=str(error))
        store.finish_setup_run(
            run_id,
            success=False,
            installed_ollama=installed_ollama,
            started_ollama=started_ollama,
            pulled_models=pulled_models,
            failed_models=failed_models,
        )
        raise HTTPException(status_code=503, detail=str(error))


@app.post("/setup/prerequisites/stream")
async def run_setup_prerequisites_stream(request: SetupPrerequisitesRequest):
    if not request.approved:
        raise HTTPException(
            status_code=400,
            detail="User approval is required before running setup.",
        )

    store = get_knowledge_store()
    run_id = store.create_setup_run(request.models)

    async def _event_stream():
        def _sse(payload: Dict[str, Any]) -> str:
            return f"data: {json.dumps(payload)}\n\n"

        installed_ollama = False
        started_ollama = False
        pulled_models: List[str] = []
        failed_models: List[Dict[str, str]] = []
        finished = False

        def emit(
            event_type: str,
            *,
            message: Optional[str] = None,
            model_name: Optional[str] = None,
            payload: Optional[Dict[str, Any]] = None,
        ) -> str:
            merged_payload: Dict[str, Any] = {"type": event_type, "run_id": run_id}
            if payload:
                merged_payload.update(payload)

            store.append_setup_run_event(
                run_id,
                event_type,
                message=message,
                model_name=model_name,
                payload=payload,
            )
            return _sse(merged_payload)

        try:
            yield emit("setup_start", payload={"requested_models": request.models or []})

            status = await build_setup_prerequisites_status()
            yield emit("status", payload={"status": status})

            if not status["ollama_installed"]:
                yield emit("install_start", payload={"platform": sys.platform})
                install_success, install_message = await install_ollama_macos()
                if not install_success:
                    raise RuntimeError(install_message)
                installed_ollama = True
                yield emit("install_done")
                status = await build_setup_prerequisites_status()
                yield emit("status", payload={"status": status})
                if not status["ollama_installed"]:
                    raise RuntimeError(
                        "Ollama installation completed but Ollama is still not detected."
                    )

            if not status["ollama_running"]:
                yield emit("start_ollama")
                started_ollama = try_start_ollama()
                if not started_ollama:
                    raise RuntimeError(
                        "Could not start Ollama automatically. Start Ollama and try again."
                    )
                if not await wait_for_ollama_ready():
                    raise RuntimeError(
                        "Ollama was started but did not become ready in time."
                    )
                yield emit("start_ollama_done")
                status = await build_setup_prerequisites_status()
                yield emit("status", payload={"status": status})

            target_models = resolve_target_models(status, request.models)
            yield emit("target_models", payload={"target_models": target_models})

            for model_name in target_models:
                yield emit("pull_start", model_name=model_name, payload={"model": model_name})
                try:
                    async for progress in stream_pull_ollama_model_progress(model_name):
                        yield emit(
                            "pull_progress",
                            model_name=model_name,
                            payload={
                                "model": model_name,
                                "status": progress.get("status"),
                                "completed": progress.get("completed"),
                                "total": progress.get("total"),
                            },
                        )
                    pulled_models.append(model_name)
                    yield emit("pull_done", model_name=model_name, payload={"model": model_name})
                except Exception as error:
                    failure_message = str(error)
                    failed_models.append({"model": model_name, "error": failure_message})
                    yield emit(
                        "pull_error",
                        model_name=model_name,
                        message=failure_message,
                        payload={
                            "model": model_name,
                            "error": failure_message,
                        },
                    )

            final_status = await build_setup_prerequisites_status()
            ready = final_status["ready"] and not failed_models
            store.finish_setup_run(
                run_id,
                success=ready,
                installed_ollama=installed_ollama,
                started_ollama=started_ollama,
                pulled_models=pulled_models,
                failed_models=failed_models,
            )
            finished = True
            yield emit(
                "complete",
                payload={
                    "approved": True,
                    "installed_ollama": installed_ollama,
                    "started_ollama": started_ollama,
                    "pulled_models": pulled_models,
                    "failed_models": failed_models,
                    "status": final_status,
                    "ready": ready,
                },
            )
        except Exception as error:
            if not finished:
                store.finish_setup_run(
                    run_id,
                    success=False,
                    installed_ollama=installed_ollama,
                    started_ollama=started_ollama,
                    pulled_models=pulled_models,
                    failed_models=failed_models,
                )
            yield emit("error", message=str(error), payload={"message": str(error)})

    return StreamingResponse(_event_stream(), media_type="text/event-stream")


@app.get("/setup/history")
async def get_setup_history(limit: int = Query(default=20, ge=1, le=200)):
    store = get_knowledge_store()
    return {"runs": store.list_setup_runs(limit)}


@app.get("/weather/status")
async def get_weather_status():
    return get_weather_status_payload()


@app.get("/weather/settings")
async def get_weather_settings():
    store = get_knowledge_store()
    return store.get_weather_settings()


@app.put("/weather/settings")
async def update_weather_settings(request: WeatherSettingsUpdateRequest):
    mode = normalize_weather_mode(request.mode)
    country = normalize_country_code(request.country)
    city = request.city.strip()
    state = request.state.strip() if request.state else None
    if not city:
        raise HTTPException(status_code=400, detail="City is required")

    try:
        matches = await resolve_openweather_location(
            city=city,
            state=state,
            country=country,
        )
    except WeatherConfigError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except WeatherAuthError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Location lookup failed: {error}")

    if not matches:
        raise HTTPException(status_code=404, detail="Location not found")

    best = matches[0]
    settings = get_knowledge_store().set_weather_settings(
        mode=mode,
        city=str(best.get("name") or city),
        state=best.get("state"),
        country=str(best.get("country") or country),
        lat=float(best["lat"]),
        lon=float(best["lon"]),
        cache_ttl_minutes=WEATHER_CACHE_TTL_MINUTES,
    )
    return settings


@app.get("/weather/resolve-location")
async def resolve_weather_location(
    city: str = Query(..., min_length=1),
    state: Optional[str] = Query(default=None),
    country: str = Query(default="US", min_length=2, max_length=3),
):
    try:
        matches = await resolve_openweather_location(
            city=city,
            state=state,
            country=normalize_country_code(country),
        )
    except WeatherConfigError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except WeatherAuthError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Location lookup failed: {error}")

    return {"results": matches}


@app.get("/weather/current")
async def get_current_weather(
    lat: float = Query(..., ge=-90.0, le=90.0),
    lon: float = Query(..., ge=-180.0, le=180.0),
    force_refresh: bool = Query(default=False),
):
    ttl = int(get_knowledge_store().get_weather_settings()["cache_ttl_minutes"])
    try:
        payload = await get_or_refresh_current_weather(
            lat=lat,
            lon=lon,
            force_refresh=force_refresh,
            ttl_minutes=ttl,
        )
    except WeatherConfigError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except WeatherAuthError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Current weather unavailable: {error}")

    return {
        "lat": lat,
        "lon": lon,
        "fetched_at": payload.get("fetched_at"),
        "source": payload.get("source"),
        "stale": bool(payload.get("stale")),
        "data": payload.get("data"),
    }


@app.get("/weather/forecast")
async def get_forecast_weather(
    lat: float = Query(..., ge=-90.0, le=90.0),
    lon: float = Query(..., ge=-180.0, le=180.0),
    force_refresh: bool = Query(default=False),
):
    ttl = int(get_knowledge_store().get_weather_settings()["cache_ttl_minutes"])
    try:
        payload = await get_or_refresh_forecast_weather(
            lat=lat,
            lon=lon,
            force_refresh=force_refresh,
            ttl_minutes=ttl,
        )
    except WeatherConfigError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except WeatherAuthError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Forecast unavailable: {error}")

    return {
        "lat": lat,
        "lon": lon,
        "fetched_at": payload.get("fetched_at"),
        "source": payload.get("source"),
        "stale": bool(payload.get("stale")),
        "forecast_3h": payload.get("forecast_3h"),
        "forecast_daily": payload.get("forecast_daily"),
    }


@app.get("/weather/alerts")
async def get_weather_alerts(
    lat: float = Query(..., ge=-90.0, le=90.0),
    lon: float = Query(..., ge=-180.0, le=180.0),
    force_refresh: bool = Query(default=False),
):
    ttl = int(get_knowledge_store().get_weather_settings()["cache_ttl_minutes"])
    try:
        payload = await get_or_refresh_alert_weather(
            lat=lat,
            lon=lon,
            force_refresh=force_refresh,
            ttl_minutes=ttl,
        )
    except WeatherConfigError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except WeatherAuthError as error:
        raise HTTPException(status_code=503, detail=str(error))
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Alerts unavailable: {error}")

    return {
        "lat": lat,
        "lon": lon,
        "fetched_at": payload.get("fetched_at"),
        "source": payload.get("source"),
        "stale": bool(payload.get("stale")),
        "alerts": payload.get("alerts") or [],
    }


@app.post("/weather/refresh")
async def refresh_weather():
    status = get_weather_status_payload()
    if not status["enabled"] and status.get("reason") == "missing_api_key":
        raise HTTPException(status_code=503, detail="OpenWeather API key is not configured.")

    store = get_knowledge_store()
    settings = store.get_weather_settings()
    location = settings.get("location")
    if not location:
        raise HTTPException(status_code=400, detail="Weather location is not configured.")

    try:
        bundle = await refresh_weather_bundle(
            lat=float(location["lat"]),
            lon=float(location["lon"]),
            force_refresh=True,
            ttl_minutes=int(settings.get("cache_ttl_minutes") or WEATHER_CACHE_TTL_MINUTES),
        )
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Weather refresh failed: {error}")

    mode = normalize_weather_mode(settings.get("mode"))
    coherence = calculate_coherence_scores(
        mode=mode,
        current=bundle.get("current"),
        forecast_points=bundle.get("forecast_3h") or [],
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
    )
    generated = generate_predictions(
        mode=mode,
        forecast_daily=bundle.get("forecast_daily") or [],
    )
    saved_predictions = store.save_predictions(
        mode=mode,
        integrity=coherence["integrity"],
        resilience=coherence["resilience"],
        meaning=coherence["meaning"],
        cci_score=coherence["cci"],
        predictions=generated,
    )
    evaluated_count = evaluate_mature_predictions()
    insights = build_weather_insights(
        mode=mode,
        coherence=coherence,
        current=bundle.get("current"),
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
    )

    return {
        "location": location,
        "mode": mode,
        "source": bundle.get("source"),
        "stale": bool(bundle.get("stale")),
        "warning": bundle.get("warning"),
        "fetched_at": bundle.get("last_refresh_ts"),
        "coherence": coherence,
        "predictions": saved_predictions,
        "insights": insights,
        "evaluated_predictions": evaluated_count,
    }


@app.get("/weather/dashboard")
async def get_weather_dashboard():
    status = get_weather_status_payload()
    if not status["enabled"] and not status.get("has_cached_data"):
        raise HTTPException(
            status_code=503,
            detail=f"Weather is unavailable: {status.get('reason') or 'disabled'}",
        )

    store = get_knowledge_store()
    settings = store.get_weather_settings()
    location = settings.get("location")
    if not location:
        raise HTTPException(status_code=400, detail="Weather location is not configured.")

    try:
        bundle = await refresh_weather_bundle(
            lat=float(location["lat"]),
            lon=float(location["lon"]),
            force_refresh=False,
            ttl_minutes=int(settings.get("cache_ttl_minutes") or WEATHER_CACHE_TTL_MINUTES),
        )
    except Exception as error:
        raise HTTPException(status_code=503, detail=f"Weather dashboard unavailable: {error}")

    mode = normalize_weather_mode(settings.get("mode"))
    coherence = calculate_coherence_scores(
        mode=mode,
        current=bundle.get("current"),
        forecast_points=bundle.get("forecast_3h") or [],
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
    )
    today = now_utc().date().isoformat()
    saved_predictions = store.list_predictions(mode=mode, date_from=today, limit=14)
    if not saved_predictions:
        generated = generate_predictions(
            mode=mode,
            forecast_daily=bundle.get("forecast_daily") or [],
        )
        saved_predictions = [
            {
                "id": f"preview-{index}",
                "created_at": str(int(time.time())),
                "mode": mode,
                "target_date": item["target_date"],
                "integrity": coherence["integrity"],
                "resilience": coherence["resilience"],
                "meaning": coherence["meaning"],
                "cci_score": coherence["cci"],
                "probability": item["probability"],
                "prediction_text": item["prediction_text"],
                "factors": item.get("factors") or {},
            }
            for index, item in enumerate(generated)
        ]

    insights = build_weather_insights(
        mode=mode,
        coherence=coherence,
        current=bundle.get("current"),
        forecast_daily=bundle.get("forecast_daily") or [],
        alerts=bundle.get("alerts") or [],
    )

    return {
        "location": location,
        "mode": mode,
        "current": bundle.get("current"),
        "forecast_daily": bundle.get("forecast_daily") or [],
        "forecast_3h": bundle.get("forecast_3h") or [],
        "alerts": bundle.get("alerts") or [],
        "coherence": coherence,
        "predictions": saved_predictions,
        "insights": insights,
        "cache_age_seconds": cache_age_seconds(bundle.get("last_refresh_ts")),
        "source": bundle.get("source"),
        "stale": bool(bundle.get("stale")),
        "warning": bundle.get("warning"),
    }


@app.get("/settings/models")
async def get_model_settings():
    store = get_knowledge_store()
    configured_models = store.get_model_names()

    try:
        available_models = await fetch_ollama_model_names()
        ollama_connected = True
    except Exception as error:
        print(
            f"Model list warning: {type(error).__name__}: {error}",
            file=sys.stderr,
        )
        available_models = []
        ollama_connected = False

    return {
        "configured_models": configured_models,
        "available_models": available_models,
        "ollama_connected": ollama_connected,
    }


@app.get("/settings/profile")
async def get_profile_settings():
    store = get_knowledge_store()
    return {"profile": store.get_assistant_profile()}


@app.put("/settings/profile")
async def update_profile_settings(request: ProfileSettingsUpdateRequest):
    store = get_knowledge_store()
    profile = store.set_assistant_profile(request.profile)
    return {"profile": profile}


@app.put("/settings/models")
async def update_model_settings(request: ModelSettingsUpdateRequest):
    next_config = {
        "lite": request.lite.strip(),
        "general": request.general.strip(),
        "deep": request.deep.strip(),
    }
    for profile_key in MODEL_PROFILE_KEYS:
        if not next_config[profile_key]:
            raise HTTPException(
                status_code=400,
                detail=f"{profile_key} model is required",
            )

    available_models: List[str] = []
    ollama_connected = True
    try:
        available_models = await fetch_ollama_model_names()
    except Exception as error:
        print(
            f"Model validation warning: {type(error).__name__}: {error}",
            file=sys.stderr,
        )
        ollama_connected = False

    if available_models:
        available_set = {
            normalize_ollama_model_name(model_name)
            for model_name in available_models
            if model_name.strip()
        }
        invalid_keys = [
            profile_key
            for profile_key in MODEL_PROFILE_KEYS
            if normalize_ollama_model_name(next_config[profile_key]) not in available_set
        ]
        if invalid_keys:
            raise HTTPException(
                status_code=400,
                detail=f"Selected model not found in Ollama for: {', '.join(invalid_keys)}",
            )

    store = get_knowledge_store()
    configured = store.set_model_names(
        lite=next_config["lite"],
        general=next_config["general"],
        deep=next_config["deep"],
    )

    return {
        "configured_models": configured,
        "available_models": available_models,
        "ollama_connected": ollama_connected,
    }


@app.get("/health")
async def health_check():
    """Health check endpoint that verifies FastAPI, Ollama, and knowledge DB connectivity."""
    health_status: Dict[str, Any] = {
        "status": "ok",
        "backend": "running",
    }

    try:
        get_knowledge_store()
        health_status["knowledge_db"] = "ready"
    except Exception as e:
        health_status["status"] = "degraded"
        health_status["knowledge_db"] = "error"
        health_status["message"] = f"Knowledge DB init failed: {type(e).__name__}"

    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            response = await client.get(f"{OLLAMA_BASE_URL}/api/version")
            if response.status_code == 200:
                health_status["ollama"] = "connected"
            else:
                health_status["status"] = "degraded"
                health_status["ollama"] = "unreachable"
                health_status["message"] = "Ollama is not responding correctly"
    except (httpx.ConnectError, httpx.TimeoutException):
        health_status["status"] = "degraded"
        health_status["ollama"] = "unreachable"
        health_status["message"] = "Cannot connect to Ollama at http://localhost:11434"
    except Exception as e:
        health_status["status"] = "degraded"
        health_status["ollama"] = "error"
        health_status["message"] = f"Error checking Ollama: {type(e).__name__}"

    return health_status
